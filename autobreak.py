#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Date    : 2017-08-10 01:32:19
# @Author  : Tural Aksel (turalaksel@gmail.com)
# @Link    : http://example.org
# @Version : $Id$

from tqdm import tqdm
from shutil import copyfile

import numpy as np
import random
import os
import sys
import argparse
import utilities
import glob
import pandas
import openpyxl
import matplotlib

import origamidesign

matplotlib.use('TkAgg')


class OligoBreakSolution:
    def __init__(self):
        self.breaks       = None
        self.edges        = None
        self.dsDNA_length = 0

    def get_cvs_rows(self):
        '''
        Prepare cvs writer object

        COLUMNS
        1.  Oligo key
        2.  Oligo Group key
        3.  Break 1 key
        4.  Break 1 type
        5.  Break 2 key
        6.  Break 2 type
        7.  Length
        8.  Sequence
        9.  Edge weight
        10.  ProbFolding
        11.  Log-Probfolding
        12.  Tf
        13. maxTm
        14. maxSeq
        15. has14
        16. dGtotal
        17. dGintrin
        18. dGloop
        19. dGconc
        '''

        # Initialize rows
        cvs_writer_rows = []
        for edge in self.edges[:-1]:
            cvs_writer_rows.append(edge.get_cvs_row_object())

        return cvs_writer_rows

    def calculate_dsDNA_length(self):
        self.dsDNA_length = 0
        for edge in self.edges:
            self.dsDNA_length += sum(edge.dsDNA_length_list)

        return self.dsDNA_length

    def break_oligo_solution(self):
        '''
        Break oligo solution
        '''
        for current_break in self.breaks[:-1]:
            current_break.break_cadnano()

    def apply_temp_neighbor_constraints(self):
        '''
        Apply neighbor constraints
        '''
        for new_break in self.breaks[:-1]:

            # Get neighbor break
            if new_break.neighbor_break:
                neighbor_break                 = new_break.neighbor_break
                neighbor_break.dont_break_temp = True

    def print_solution(self):
        '''
        Print break solution
        '''
        if self.breaks:
            tqdm.write('Break path:\n'+'->\n'.join(["(%3d.%3d.%3d)" %
                       (current_break.key) for current_break in self.breaks]))
            tqdm.write('Edge length/weight: '+'->'.join(['(%d / %.1e)' %
                       (edge.edge_length, edge.edge_weight) for edge in self.edges[:-1]]))

    def reset_temp_neighbor_constraints(self):
        '''
        Reset neighbor constraints
        '''
        for new_break in self.breaks[:-1]:

            # Get neighbor break
            if new_break.neighbor_break:
                neighbor_break                 = new_break.neighbor_break
                neighbor_break.dont_break_temp = False

    def calculate_self_penalty(self):
        '''
        Calculate self penalty score
        '''
        self.self_penalty = 0
        for new_break in self.breaks[:-1]:

            # Get neighbor break
            if new_break.neighbor_break:

                neighbor_break     = new_break.neighbor_break

                if neighbor_break and neighbor_break in self.breaks:

                    # Update total penalty score
                    self.self_penalty += 1

        # Divide penalty by 2
        self.self_penalty = int(1.0*self.self_penalty/2)

    def initialize(self):
        '''
        Prepare the lists
        '''
        self.breaks = [break_path.break_node for break_path in self.break_paths[::-1]]
        self.edges  = [break_path.break_edge for break_path in self.break_paths[::-1]]
        self.scores = [break_path.score for break_path in self.break_paths[::-1]]

    def is_identical(self, other_solution, max_index=None):
        '''
        Compare the baths between two solutions
        '''
        # Determine maximum index
        max_i = len(self.breaks)-1
        if max_index:
            max_i = max_index

        # Identical variable
        identical = True

        # Breaks for the current path
        current_breaks = self.breaks[:max_i]

        # Other breaks
        other_breaks   = other_solution.breaks[:max_i]

        if not len(current_breaks) == len(other_breaks):
            return False

        # Pairwise comparison of elements
        for i in range(max_i):
            identical *= (current_breaks[i] == other_breaks[i])

        return identical


class GroupBreaksolution:

    def __init__(self):
        '''
        Break solutions for oligo group
        '''

        self.break_solutions    = None
        self.total_score        = 0
        self.total_penalty      = 0
        self.total_dsDNA_length = 0
        self.complete           = True

    def get_cvs_rows(self):
        cvs_writer_rows = []
        for key in self.break_solutions:
            # Get break solution
            break_solution    = self.break_solutions[key]

            # If the solution doesnt exist move to next break solution
            if not break_solution:
                continue

            # Extend the list with the row objects
            cvs_writer_rows.extend(break_solution.get_cvs_rows())

        return cvs_writer_rows

    def break_group_solution(self):
        '''
        Break group solution
        '''
        for key in self.break_solutions:
            # Get break solution
            break_solution    = self.break_solutions[key]

            # If the solution doesnt exist move to next break solution
            if not break_solution:
                continue

            # Perform break
            break_solution.break_oligo_solution()

    def print_solution(self):
        '''
        Print group solution
        '''
        if self.break_solutions:
            tqdm.write('Complete:%-5s TotalScore:%-5.2f - TotalCrossoverPenalty:%-3d' %
                       (self.complete, self.total_score, self.total_penalty))
            # Print the solutions
            for oligo_key in self.break_solutions:
                # Print solution for oligo
                tqdm.write('Solution for oligo: (%d,%d,%d)' % oligo_key)
                if self.break_solutions[oligo_key]:
                    self.break_solutions[oligo_key].print_solution()

    def calculate_penalty(self, verbose=False):
        '''
        Calculate penalty for the oligo group solution
        '''
        self.total_score        = 0
        self.total_penalty      = 0
        self.total_dsDNA_length = 0
        self.complete           = True

        # Iterate over each solution
        for key in self.break_solutions:
            # Get break solution
            break_solution    = self.break_solutions[key]

            # If the solution doesnt exist move to next break solution
            if not break_solution:
                if verbose:
                    tqdm.write('SOLUTION DOESNT EXIST for oligo (%d,%d,%d)' % (key[0], key[1], key[2]))
                self.complete = False
                continue

            # Update total score
            self.total_score += break_solution.score

            # Update totel dsDNA length
            self.total_dsDNA_length += break_solution.calculate_dsDNA_length()

            # Initialize the bad break list
            break_solution.bad_list = []

            # Iterate over breaks in breaks list
            for new_break in break_solution.breaks[:-1]:

                # Get neighbor key
                if new_break.neighbor_break:
                    neighbor_break     = new_break.neighbor_break
                    neighbor_oligo_key = neighbor_break.oligo.key

                    # Get the neighbor information
                    if(neighbor_oligo_key in self.break_solutions and self.break_solutions[neighbor_oligo_key] and
                       neighbor_break in self.break_solutions[neighbor_oligo_key].breaks):
                        break_solution.bad_list.append(new_break)

                        # Update total penalty score
                        self.total_penalty += 1

        # Divide penalty score by 2
        self.total_penalty = int(self.total_penalty/2)


class CompleteBreakSolution:
    '''
    Complete break solution class
    '''

    def __init__(self):

        self.temperature_celcius = None
        self.group_solutions     = {}
        self.total_prob          = 1
        self.total_score         = 0
        self.total_norm_score    = 0
        self.total_penalty       = 0
        self.total_dsDNA_length  = 0
        self.sequence_offset     = 0
        self.corrected_offset    = 0
        self.complete            = True
        self.cvs_writer_rows     = []

        '''
        COLUMNS
        1.  Oligo key
        2.  Oligo Group key
        3.  Break 1 key
        4.  Break 1 type
        5.  Break 1 location
        5.  Break 2 key
        6.  Break 2 type
        7.  Break 2 location
        8.  Length
        9.  Sequence
        10. Edge weight
        11. ProbFolding
        12. Log-Probfolding
        13. Tf
        14. maxTm
        15. maxSeq
        16. has14
        17. dGtotal
        18. dGintrin
        19. dGloop
        20. dGconc
        '''
        self.cvs_header = ['OligoKey',
                           'OligoGroupkey',
                           'StartBreakKey',
                           'StartBreakType',
                           'StartBreakLocation',
                           'EndBreakKey',
                           'EndBreakType',
                           'EndBreakLocation',
                           'Length',
                           'Sequence',
                           'EdgeWeight',
                           'ProbFold',
                           'LogProbFold',
                           'Tf',
                           'maxTm',
                           'maxSeqLength',
                           'Has14',
                           'dGtotal',
                           'dGintrin',
                           'dGLoop',
                           'dGconc']

    def calculate_total_score(self):
        '''
        Calculate total score for the best solutions
        '''
        self.total_prob         = 1.0
        self.total_score        = 0
        self.total_penalty      = 0
        self.total_dsDNA_length = 0
        self.complete           = True

        for key in self.group_solutions:
            # Break group solution
            if self.group_solutions[key]:
                self.total_prob         *= np.exp(self.group_solutions[key].total_score)
                self.total_score        += self.group_solutions[key].total_score
                self.total_penalty      += self.group_solutions[key].total_penalty
                self.total_dsDNA_length += self.group_solutions[key].total_dsDNA_length
                self.complete           *= self.group_solutions[key].complete
            else:
                self.complete       = False

        # Determine normalized score
        self.total_norm_score = 1.0*self.total_score/self.total_dsDNA_length

    def get_cvs_rows(self):
        '''
        Get cvs writer rows
        '''
        cvs_writer_rows = []
        for key in self.group_solutions:
            # Check if the solution exists
            if self.group_solutions[key]:
                cvs_writer_rows.extend(self.group_solutions[key].get_cvs_rows())

        return cvs_writer_rows

    def get_summary_rows(self):
        '''
        Get summary rows
        '''
        summary_rows = [['Temperature', self.temperature_celcius],
                        ['TotalProb', self.total_prob],
                        ['TotalScore', self.total_score],
                        ['TotalNormScore', self.total_norm_score],
                        ['TotalPenalty', self.total_penalty],
                        ['Complete', int(self.complete)],
                        ['SequenceOffset', self.sequence_offset],
                        ['CorrectedOffset', self.corrected_offset]]

        return summary_rows

    def export_staples(self, filename):
        '''
        Export the staples and its scores into an excel file
        '''
        # Get summary rows
        summary_rows = np.array(self.get_summary_rows())

        # Get the cvs rows
        cvs_rows = np.array(self.get_cvs_rows())

        # If the data arrays are empty, quit
        if len(summary_rows) == 0 or len(cvs_rows) == 0:
            return

        # Load workbook
        book = openpyxl.load_workbook(filename)
        writer      = pandas.ExcelWriter(filename, engine='openpyxl')
        writer.book = book

        # Assign sheet number
        sheet_number = self.sequence_offset

        # Create data frames
        summary_frame = pandas.DataFrame(summary_rows)

        # Create staples frames
        staples_frame  = pandas.DataFrame(cvs_rows)

        # Write summary data
        summary_frame.to_excel(writer, sheet_name=str(sheet_number), header=None, index=False)

        # Write staples data
        staples_frame.to_excel(writer, sheet_name=str(sheet_number), header=self.cvs_header, index=False, startrow=10)

        # Save writer and close
        writer.save()
        writer.close()


class AutoStaple:
    def __init__(self):
        '''
        Auto staple class
        '''
        self.origami = None

    def decorate(self):
        '''
        Decorate the design
        '''
        self.part.potentialCrossoverMap(0)


class AutoBreak:
    def __init__(self):

        # Cadnano parameters
        self.origami                         = None
        self.json_input                      = None

        # Constraints
        self.MAX_OLIGO_LENGTH                = 60
        self.UPPER_BOUND                     = 60
        self.LOWER_BOUND                     = 21

        # Local and global solutions
        self.NUM_OLIGO_SOLUTIONS             = 100
        self.NUM_GLOBAL_SOLUTIONS            = 1

        # Score parameters
        self.total_score                     = 0
        self.total_penalty                   = 0

        # k-shortest path parameter
        self.k_select                        = 'best'

        # Break rule
        self.break_rule                      = ['xstap', 'all3']

        # Maximum number of breaks allowed per oligo for autobreak
        self.max_num_breaks                  = 10000000

        # Readonly setting
        self.readonly                        = False

        # Optimization parameters
        self.optim_shuffle_oligos            = True
        self.optim_pick_method               = 'best'

        # Output file
        self.json_legacy_output             = 'out_legacy.json'
        self.json_cn25_output               = 'out_cn25.json'
        self.output_directory                = '.'

        # Optimization function
        self.optim_args                      = [['dG:50']]

        # Maxseq parameter
        self.optim_maxseq_length             = 14

        # Length gaussian parameters
        self.optim_length_mean               = 45
        self.optim_length_tolerance          = 5

        # Tm gaussian parameters
        self.optim_Tm_mean                   = 60
        self.optim_Tm_tolerance              = 5

        # Temperature parameters for dG optimization
        self.optim_temperature_celcius       = 50
        self.optim_temperature_kelvin        = self.optim_temperature_celcius+273.15

        # Max sequence gaussian parameters
        self.optim_maxseq_mean               = 14
        self.optim_maxseq_tolerance          = 2

        # Set optimization score functions
        self.optim_score_functions           = ['sum']

        # Structure factor
        self.optim_structure_factor         = 1.0

        # Set function dictionary
        self.optim_funcs_dict                = {'14': self._optimize_14,
                                                '16': self._optimize_16,
                                                'Tm': self._optimize_Tm,
                                                'dG': self._optimize_dG,
                                                'structure': self._optimize_structure,
                                                'length': self._optimize_length,
                                                'maxseq': self._optimize_maxseq,
                                                'glength': self._gauss_length,
                                                'gmaxseq': self._gauss_maxseq,
                                                'gTm': self._gauss_Tm,
                                                'llength': self._log_length,
                                                'lmaxseq': self._log_maxseq,
                                                'lTm': self._log_Tm}
        # Set optimization params
        self.optim_params_dict              = {'14': [],
                                               '16': [],
                                               'Tm': [],
                                               'dG': [self.optim_temperature_celcius],
                                               'structure': [self.optim_structure_factor],
                                               'length': [],
                                               'maxseq': [self.optim_maxseq_length],
                                               'glength': [self.optim_length_mean, self.optim_length_tolerance],
                                               'gmaxseq': [self.optim_maxseq_mean, self.optim_maxseq_tolerance],
                                               'gTm': [self.optim_Tm_mean, self.optim_Tm_tolerance],
                                               'llength': [self.optim_length_mean, self.optim_length_tolerance],
                                               'lmaxseq': [self.optim_maxseq_mean, self.optim_maxseq_tolerance],
                                               'lTm': [self.optim_Tm_mean, self.optim_Tm_tolerance]}

        # Verbose output
        self.verbose_output                 = False

        # Solutions containers
        self.complete_solutions             = {}
        self.best_complete_solution         = None

        # Sequence parameter
        self.best_sequence_offset           = 0

        # Permutation parameter
        self.permute_sequence               = False

        # Excel file that stores the results
        self.results_excel_file             = None
        self.autobreak_excel_file           = None
        self.summary_excel_file             = None

    def get_results_summary(self):
        '''
        Get results summary
        '''
        self.results_summary = []

        # Print the scores
        for complete_solution in self.sorted_complete_solutions:
            self.results_summary.append([complete_solution.sequence_offset,
                                         complete_solution.corrected_offset,
                                         complete_solution.total_prob,
                                         complete_solution.total_score,
                                         complete_solution.total_norm_score,
                                         complete_solution.total_penalty])

        return self.results_summary

    def write_results_summary(self):
        '''
        Write results summary
        '''

        # Get results summary
        results_summary = np.array(self.get_results_summary())

        # If the array is empty, quit
        if len(results_summary) == 0:
            sys.exit('SOLUTION DOESNT EXIST')

        # Create writer
        writer      = pandas.ExcelWriter(self.summary_excel_file, engine='openpyxl')

        # Create data frames
        summary_frame  = pandas.DataFrame(results_summary)

        # Create summary header
        summary_header = ['SequenceOffset', 'CorrectedOffset', 'TotalProb', 'TotalScore', 'TotalNormScore', 'TotalPenalty']

        # Write summary data
        summary_frame.to_excel(writer, sheet_name='summary', header=summary_header, index=False)

        # Save writer and close
        writer.save()
        writer.close()

    def write_results(self, sequence_offset=0):
        '''
        Write Solution results to a sheet in results excel file
        '''
        if sequence_offset in self.complete_solutions:
            self.complete_solutions[sequence_offset].export_staples(self.results_excel_file)

    def write_best_result(self):
        '''
        Write best result
        '''
        if not self.write_all_results:
            self.best_complete_solution.export_staples(self.results_excel_file)

    def create_results_excel_file(self):
        '''
        Create resuls excel file
        '''
        wb = openpyxl.Workbook()
        wb.save(self.results_excel_file)

    def set_lower_bound(self, min_length=21):
        '''
        Set lower bound
        '''
        self.LOWER_BOUND = min_length

    def set_readonly(self, read_only=False):
        '''
        Set readonly parameter
        '''
        self.readonly = read_only

    def set_write_all_results(self, write_all_results=False):
        '''
        Set write all results
        '''
        self.write_all_results = write_all_results

    def set_temperature_parameter(self):
        '''
        Set temperature parameters
        '''
        self.optim_temperature_celcius = self.optim_params_dict['dG'][0]
        self.optim_temperature_kelvin  = self.optim_temperature_celcius + 273.15

    def set_permute_sequence(self, permute=False):
        '''
        Set permute sequence
        '''
        self.permute_sequence = permute

    def color_oligos_by_folding_prob(self):
        '''
        Color oligos by folding prob
        '''
        for oligo in self.origami.oligos['staple']:
            oligo.color_by_folding_prob()

    def color_oligos_by_Tf(self):
        '''
        Color oligos by folding prob
        '''
        for oligo in self.origami.oligos['staple']:
            oligo.color_by_Tf()

    def preprocess_optim_params(self):
        '''
        Preprocess optimization parameters


        RULE1. If there is no cross in break rule,
               make oligo solution and global solution number 1
               since Optimization yields the best result

        RUL2. If oligos are not shuffled, make num oligo solutions
              and global solutions 1

        '''

        # RULE 1
        if 'xscaf' not in self.break_rule and 'xstap' not in self.break_rule:
            self.NUM_OLIGO_SOLUTIONS  = 1
            self.NUM_GLOBAL_SOLUTIONS = 1

        # RULE 2
        if not self.optim_shuffle_oligos:
            self.NUM_OLIGO_SOLUTIONS  = 1
            self.NUM_GLOBAL_SOLUTIONS = 1

    def set_output_directory(self, input_filename):
        '''
        Set output directory
        '''
        # Split input file
        head, tail       = os.path.split(os.path.abspath(input_filename))
        root, ext        = os.path.splitext(tail)

        # List existing output directories
        potential_directories = list(filter(lambda x: os.path.isdir(x),
                                            glob.glob(head+'/'+root+'_autobreak_[0-9][0-9][0-9]')))

        # Get the number extensions
        number_extensions = [int(x[-3:]) for x in potential_directories]

        # Get the counter
        output_counter = 1
        if len(number_extensions) > 0:
            output_counter = max(number_extensions)+1

        self.output_directory = head+'/'+root+"_autobreak_%03d" % (output_counter)

        # Make directory
        os.mkdir(self.output_directory)

        # Copy input file to output directory
        copyfile(input_filename, self.output_directory+'/'+tail)

    def copy_sequence_file(self):
        '''
        Copy sequence file to output directory
        '''
        if self.origami.sequence_file and os.path.isfile(self.origami.sequence_file):
            head, tail = os.path.split(self.origami.sequence_file)
            copyfile(self.origami.sequence_file, self.output_directory+'/'+tail)
        elif self.origami.sequence_file not in utilities.SCAFFOLD_SEQUENCES:
            f = open(self.output_directory+'/random_sequence.txt', 'w')
            f.write(self.origami.scaffold_sequence)
            f.close()

    def write_input_args(self):
        '''
        Write input arguments in a txt file
        '''
        self.args_file = self.output_directory+'/args.txt'
        f = open(self.args_file, 'w')
        for key in self.args_dict:
            f.write('%-8s: %-8s\n' % (key, self.args_dict[key]))

        f.close()

    def define_output_files(self):
        '''
        Define json output
        '''

        # Split input file
        head, tail       = os.path.split(self.origami.json_input)
        root, ext        = os.path.splitext(tail)

        # Output files
        self.json_start_output  = self.output_directory+'/'+root+'_start.json'

        self.json_legacy_output = self.output_directory+'/'+root+'_autobreak_legacy.json'
        self.json_cn25_output   = self.output_directory+'/'+root+'_autobreak_cn25.json'

        # Results excel file
        self.results_excel_file   = self.output_directory+'/'+root+'.xlsx'
        self.autobreak_excel_file = self.output_directory+'/'+root+'_autobreak.xlsx'
        self.summary_excel_file   = self.output_directory+'/'+root+'_summary.xlsx'

    def write_part_to_json(self, filename, legacy_option=True):
        '''
        Write part to json
        '''
        self.origami.doc.writeToFile(filename, legacy=legacy_option)

    def write_start_part_to_json(self):
        '''
        Write the starting point for autobreak
        '''
        self.origami.doc.writeToFile(self.json_start_output, legacy=True)

    def write_final_part_to_json(self):
        '''
        Write cadnano part to json
        '''
        self.origami.doc.writeToFile(self.json_legacy_output, legacy=True)
        self.origami.doc.writeToFile(self.json_cn25_output,   legacy=False)

    def set_pick_method(self, pick_method='random'):
        '''
        Set solution picking method
        '''
        self.optim_pick_method = pick_method

    def set_maximum_breaks(self, maximum_num_breaks=100000):
        '''
        Set maximum number of breaks parameter
        '''
        self.max_num_breaks = maximum_num_breaks

    def set_oligo_shuffle_parameter(self, oligo_shuffle_parameter=False):
        '''
        Set oligo shuffle parameter
        '''
        self.optim_shuffle_oligos = oligo_shuffle_parameter

    def set_solution_nums(self, solutions_per_oligo=1000, global_solutions=1000):
        '''
        Set solution numbers
        '''
        self.NUM_OLIGO_SOLUTIONS  = solutions_per_oligo
        self.NUM_GLOBAL_SOLUTIONS = global_solutions

    def set_k_select(self, k_parameter='best'):
        '''
        Set k-select value
        '''
        self.k_select = k_parameter

    def set_break_rule(self, new_break_rule=['xstap', 'all2']):
        '''
        Set break rule
        '''
        self.break_rule         = [rule for rule in new_break_rule]
        self.origami.break_rule = self.break_rule

    def set_verbose_output(self, verbose=False):
        '''
        Set verbose output
        '''
        self.verbose_output = verbose

    def shift_scaffold_sequence(self, offset=0):
        '''
        Shift scaffold sequence and update autobreak graphs
        '''
        # Apply the offset and shift sequence
        self.origami.apply_sequence(offset)

        # Update strand and sequence dna
        self.origami.assign_strands_dna()
        self.origami.update_sequences_dna()

        # Treat sequence update differently
        if self.readonly:
            self.determine_initial_scores()
        else:
            # Update graph edge weights
            self.update_edge_weights()

    def permute_scaffold_sequence_autobreak(self, nitr=100):
        '''
        Permute scaffold sequence
        '''
        # Set start offset
        start_offset = self.origami.sequence_offset

        # Set final offset
        final_itr = 1
        if self.permute_sequence:
            final_itr = len(self.origami.scaffold_sequence)
            # Check the number of permutation iterations allowed
            if nitr < final_itr:
                final_itr = nitr

        for itr in tqdm(range(0, final_itr), desc='Permutation loop', leave=False,
                        dynamic_ncols=True, bar_format='{desc}: {percentage:3.2f}%|'+'{bar}'):

            # Set the current offset
            current_offset = (start_offset+itr) % self.origami.scaffolds[0].length()

            # Shift the sequence to the offset
            self.shift_scaffold_sequence(current_offset)

            # Run autobreak
            self.run_autobreak()

            # Write results
            if self.write_all_results:
                self.write_results(current_offset)

    def permute_scaffold_sequence_readonly(self, nitr=100):
        '''
        Permute scaffold sequence
        '''
        # Set start offset
        start_offset = self.origami.sequence_offset

        # Set final offset
        final_itr = 1
        if self.permute_sequence:
            final_itr = len(self.origami.scaffold_sequence)
            # Check the number of permutation iterations allowed
            if nitr < final_itr:
                final_itr = nitr

        for itr in tqdm(range(0, final_itr), desc='Permutation loop', leave=False,
                        dynamic_ncols=True, bar_format='{desc}: {percentage:3.2f}%|'+'{bar}'):

            # Set the current offset
            current_offset = (start_offset+itr) % self.origami.scaffolds[0].length()

            # Shift the sequence to the offset
            self.shift_scaffold_sequence(current_offset)

            # Determine scores and add the solution to solutions list
            self.determine_readonly_scores()

    def run_autobreak(self):
        '''
        Run basic autobreak protocol
        '''

        # Create stepwise group solutions
        self.create_stepwise_group_solutions()

        # Sort and print group solutions
        self.sort_group_solutions()

        # Combine group solutions
        self.combine_group_solutions()

    def determine_oligo_scores(self):
        '''
        Determine oligo scores from design
        '''
        # Make break-break graph
        self.origami.prepare_origami()

        # Determine initial scores
        self.determine_initial_scores()

    def create_independent_group_solutions(self):
        '''
        Create independent group solution
        No temporary neighbor constraints are imposed during run
        '''

        # Break oligos
        self.create_oligo_solutions()

        # Create group solutions
        self.combine_oligo_solutions()

    def create_stepwise_group_solutions(self):
        '''
        Main function for solution determination
        '''
        for oligo_group in tqdm(self.origami.oligo_groups, desc='Main loop ',
                                dynamic_ncols=True, bar_format='{l_bar}{bar}'):

            # Sort oligos by length
            oligo_group.sort_oligos_by_length(reverse=True)

            # Create solutions via stepwise approach
            oligo_group.create_stepwise_oligo_solutions(self.NUM_OLIGO_SOLUTIONS, self.NUM_GLOBAL_SOLUTIONS,
                                                        self.optim_pick_method, self.optim_shuffle_oligos,
                                                        verbose=self.verbose_output)

            # Remove incomplete solutions
            oligo_group.remove_incomplete_solutions()

    def update_edge_weights(self):
        '''
        Update edge weights
        '''
        for oligo in self.origami.oligos['staple']:
            # Visit each break object
            for current_break in oligo.breaks:
                # Iterate over the break edges
                for break_edge in current_break.break_edges:
                    # Update edge weights
                    break_edge.update_connection()

    def initialize(self):
        '''
        Initialize the connectivity maps
        '''

        for oligo in self.origami.oligos['staple']:
            # Check oligo length, if the length is within length limits dont break it
            if oligo.length < self.LOWER_BOUND:
                oligo.dont_break = True

            # Visit each break object
            for current_break in oligo.breaks:
                # Initialize the break edges
                current_break.break_edges = []

                # Get next break
                next_break = current_break.next_break
                # Iterate over the breaks
                while next_break:

                    # Determine break to break distance
                    break_distance = current_break.get_break_distance(next_break)

                    if break_distance >= self.LOWER_BOUND and break_distance <= self.UPPER_BOUND:
                        # Create break
                        new_edge = BreakEdge()

                        # Assign origami
                        new_edge.origami = self.origami

                        # Assign autobreak
                        new_edge.autobreak = self

                        # Assign edge length
                        new_edge.edge_length = break_distance

                        # Make the connection
                        new_edge.make_connection(current_break, next_break)

                        # Make loop edge
                        new_edge.make_loop_edge()

                        # Set edge weight
                        new_edge.edge_weight = self.optimize(new_edge)

                        # Add directed edge to current break's edges
                        current_break.break_edges.append(new_edge)

                        # Add break edge to edge map
                        self.origami.break_edge_map[current_break.key+next_break.key] = new_edge

                    # Stop criteria
                    if break_distance > self.UPPER_BOUND or next_break == current_break:
                        break

                    next_break = next_break.next_break

    def reset_temp_neighbor_constraints(self):
        '''
        Reset temporary neighbor constraints
        '''
        for oligo in self.origami.oligos['staple']:
            oligo.reset_temp_neighbor_constraints()

    def determine_initial_scores(self):
        '''
        Determine starting scores
        '''
        for oligo in self.origami.oligos['staple']:
            oligo.get_initial_score()

    def determine_readonly_scores(self):
        '''
        Determine readonly scores at a the current settings
        '''
        # Create a Complete Break Solution object
        new_complete_solution = CompleteBreakSolution()

        # Save total score and prob
        total_score        = 0
        total_prob         = 1.0
        total_dsDNA_length = 0
        for oligo in self.origami.oligos['staple']:
            # Add scores
            total_score        += oligo.initial_score
            total_prob         *= oligo.folding_prob
            total_dsDNA_length += oligo.dsDNA_length

        # Assign the scores
        new_complete_solution.total_prob         = total_prob
        new_complete_solution.total_score        = total_score
        new_complete_solution.sequence_offset    = self.origami.sequence_offset
        new_complete_solution.total_dsDNA_length = total_dsDNA_length

        # Determine normalizced score
        new_complete_solution.total_norm_score = 1.0*total_score/total_dsDNA_length
        # Add the solution to solutions list
        self.complete_solutions[self.origami.sequence_offset] = new_complete_solution

    def export_initial_scores(self):
        '''
        Export initial scores to final excel file
        '''

        # Create a dummy Complete Break Solution object
        self.final_complete_break_solution = CompleteBreakSolution()

        # Store final cvs rows
        self.final_cvs_rows = []

        # Save total score and prob
        total_score = 0
        total_prob  = 1.0
        total_dsDNA = 0.0
        for oligo in self.origami.oligos['staple']:
            self.final_cvs_rows.append(oligo.end_to_end_edge.get_cvs_row_object())

            # Add scores
            total_score += oligo.initial_score
            total_prob  *= oligo.folding_prob
            total_dsDNA += oligo.dsDNA_length

        # Determine normalizced score
        total_norm_score = 1.0*total_score/total_dsDNA

        # Prepare summary data
        self.final_summary_data = [['SequenceOffset', self.best_complete_solution.sequence_offset],
                                   ['CorrectedOffset', self.best_complete_solution.corrected_offset],
                                   ['TotalProb', total_prob],
                                   ['TotalScore', total_score],
                                   ['TotalNormScore', total_norm_score]]

        # Check if the data arrays are empty
        if len(self.final_cvs_rows) == 0:
            return

        # Results header
        cvs_header = self.final_complete_break_solution.cvs_header

        # Write the results
        writer      = pandas.ExcelWriter(self.autobreak_excel_file,
                                         engine='openpyxl')

        # Create data frames
        summary_frame = pandas.DataFrame(np.array(self.final_summary_data))

        # Create staples frames
        staples_frame  = pandas.DataFrame(np.array(self.final_cvs_rows))

        # Write summary data
        summary_frame.to_excel(writer, sheet_name=str('final'), header=None, index=False)

        # Write staples data
        staples_frame.to_excel(writer, sheet_name=str('final'), header=cvs_header, index=False, startrow=10)

        # Save writer and close
        writer.save()
        writer.close()

    def create_oligo_solutions(self):
        '''
        Break oligos
        '''
        for oligo in self.origami.oligos['staple']:
            if not oligo.dont_break:
                oligo.generate_shortest_paths(self.NUM_OLIGO_SOLUTIONS, verbose=self.verbose_output)
                oligo.remove_penalized_solutions()

    def combine_oligo_solutions(self):
        '''
        Create oligo group solutions
        '''
        for oligo_group in self.origami.oligo_groups:
            oligo_group.combine_oligo_solutions(self.NUM_GLOBAL_SOLUTIONS)

    def sort_group_solutions(self):
        '''
        Sort group solutions based on penalty score
        '''
        for oligo_group in self.origami.oligo_groups:
            oligo_group.sort_solutions()

            # If verbose output print solutions
            if self.verbose_output:
                oligo_group.print_solutions()

    def combine_group_solutions(self):
        '''
        Combine group solutions
        '''

        # Make new complete solution
        new_complete_solution = CompleteBreakSolution()
        new_complete_solution.group_solutions = {}
        new_complete_solution.sequence_offset = self.origami.sequence_offset

        # Iterate over all the oligo groups
        for oligo_group in self.origami.oligo_groups:

            # Add the best and best penalty solutions
            new_complete_solution.group_solutions[oligo_group.key]  = oligo_group.best_score_solution

        # Calculate the total score and penalty for the complete solution
        new_complete_solution.calculate_total_score()

        # Assign the best complete solution if it is complete
        if new_complete_solution.complete:
            self.complete_solutions[self.origami.sequence_offset] = new_complete_solution

    def compare_complete_solutions(self):
        '''
        Compare complete solutions
        '''
        # Sort complete solutions
        self.sorted_complete_solutions = sorted(list(self.complete_solutions.values()),
                                                key=lambda solution: solution.total_score, reverse=True)

        # Print the scores
        for complete_solution in self.sorted_complete_solutions:
            # Print total score and crossover penalty for the best solution
            tqdm.write('Complete solutions: Offset: %-5d' % (complete_solution.sequence_offset) +
                       ' - CorrectedOffset: %-5d' % (complete_solution.corrected_offset) +
                       ' - TotalProb:%-5.2f' % (complete_solution.total_prob) +
                       ' - TotalScore:%-5.2f' % (complete_solution.total_score) +
                       ' - TotalNormScore:%-5.2f' % (complete_solution.total_norm_score) +
                       ' - TotalCrossoverPenalty:%-3d' % (complete_solution.total_penalty))

        # Assign best solution
        if len(self.sorted_complete_solutions) > 0:
            self.best_complete_solution = self.sorted_complete_solutions[0]

    def correct_complete_solution_offsets(self):
        '''
        Correct complete solution offsets based on the difference between actualy start pos
        and the current start positions
        '''
        offset_difference = self.origami.get_scaffold_distance(self.origami.current_start_pos,
                                                               self.origami.sequence_start_pos)

        # Subtract offset difference from the offsets
        for key in self.complete_solutions:
            complete_solution = self.complete_solutions[key]
            complete_solution.corrected_offset = (complete_solution.sequence_offset +
                                                  offset_difference) % self.origami.scaffolds[0].length()

    def set_best_sequence_offset(self):
        '''
        Set best sequence offset
        '''
        # Best sequence offset
        if self.best_complete_solution:
            self.origami.sequence_offset = self.best_complete_solution.sequence_offset

    def break_best_complete_solution(self):
        '''
        Oligo breaking routine
        '''
        if self.best_complete_solution is None:
            sys.exit('SOLUTION DOESNT EXIST!')

        for key in self.best_complete_solution.group_solutions:
            # Break group solution
            if self.best_complete_solution.group_solutions[key]:
                self.best_complete_solution.group_solutions[key].break_group_solution()

    def set_score_func(self, func_args):
        '''
        Set optimization score functions
        '''
        self.optim_score_functions = func_args

    def set_optimization_func(self, func_args):
        '''
        Set optimization function
        '''
        self.optim_args         = func_args
        self.optim_args_funcs   = [function[0] for function in func_args]
        self.optim_args_params  = [[float(x) for x in function[1:]] for function in func_args]
        self.optimize_func_list = []

        # Set optimization function parameters
        for i in range(len(self.optim_args_funcs)):
            func   = self.optim_args_funcs[i]
            params = self.optim_args_params[i]

            # Make the optimize function
            if func in self.optim_funcs_dict:
                self.optimize_func_list.append(self.optim_funcs_dict[func])

            # Assign function parameters
            if func in self.optim_params_dict and len(params) <= len(self.optim_params_dict[func]):
                for j in range(len(params)):
                    self.optim_params_dict[func][j] = params[j]

    def optimize(self, edge):
        '''
        final optimization function
        '''
        score = 0
        # Get the score for each function type
        score_list = np.array([func(edge) for func in self.optimize_func_list])

        if 'sum' in self.optim_score_functions:
            score += np.sum(score_list)
        if len(score_list) > 1 and 'product' in self.optim_score_functions:
            score += np.product(score_list)
        return score

    def _optimize_structure(self, edge):
        '''
        Optimization function for structure
        '''
        return edge.edge_structure*self.optim_params_dict['structure'][0]

    def _optimize_dG(self, edge):
        '''
        Optimization function dG
        '''
        return edge.edge_logprob

    def _optimize_14(self, edge):
        '''
        Optimization function 14
        '''
        return edge.edge_has14

    def _optimize_16(self, edge):
        '''
        Optimization function 16
        '''
        return edge.edge_has16

    def _optimize_length(self, edge):
        '''
        Optimization function 16
        '''
        return edge.edge_length

    def _optimize_maxseq(self, edge):
        '''
        Optimization function for N
        '''
        return int(edge.edge_maxseq >= self.optim_params_dict['maxseq'][0])

    def _optimize_Tm(self, edge):
        '''
        Optimization function Tm
        '''
        return edge.edge_maxTm

    def _gauss_length(self, edge):
        '''
        Optimization function gauss length
        '''

        return np.exp(-(edge.edge_length -
                      self.optim_params_dict['glength'][0])**2/self.optim_params_dict['glength'][1]**2)

    def _gauss_Tm(self, edge):
        '''
        Optimization function gauss Tm
        '''

        return np.exp(-(edge.edge_maxTm-self.optim_params_dict['gTm'][0])**2/self.optim_params_dict['gTm'][1]**2)

    def _gauss_maxseq(self, edge):
        '''
        Optimization function gauss Tm
        '''

        return np.exp(-(edge.edge_maxseq -
                      self.optim_params_dict['gmaxseq'][0])**2/self.optim_params_dict['gmaxseq'][1]**2)

    def _log_length(self, edge):
        '''
        Optimization function gauss length
        '''

        return (-(edge.edge_length -
                self.optim_params_dict['glength'][0])**2/self.optim_params_dict['glength'][1]**2)

    def _log_Tm(self, edge):
        '''
        Optimization function gauss Tm
        '''

        return (-(edge.edge_maxTm-self.optim_params_dict['gTm'][0])**2/self.optim_params_dict['gTm'][1]**2)

    def _log_maxseq(self, edge):
        '''
        Optimization function gauss Tm
        '''

        return (-(edge.edge_maxseq -
                self.optim_params_dict['gmaxseq'][0])**2/self.optim_params_dict['gmaxseq'][1]**2)


class BreakEdge:
    def __init__(self):
        '''
        Break edge class
        '''
        self.autobreak     = None
        self.origami       = None
        self.current_break = None
        self.next_break    = None
        self.edge_weight   = None
        self.edge_length   = None
        self.edge_maxseq   = None
        self.edge_Tm       = None
        self.LOW_TM        = 60

        # state parameters
        self.active        = True
        self.valid         = True

        # Length parameters
        self.edge_has14    = None
        self.edge_num14    = None

        self.edge_has16    = None
        self.edge_num16    = None

        # Tm parameters
        self.edge_hasTm    = None
        self.edge_numTm    = None

        self.sequence_list = None

        # Loop parameter
        self.isloop        = False

    def get_cvs_row_object(self):
        '''
        Return cvs row object

        COLUMNS
        1.  Oligo key
        2.  Oligo Group key
        3.  Break 1 key
        4.  Break 1 type
        5.  Break 1 location
        6.  Break 2 key
        7.  Break 2 type
        8.  Break 2 location
        9.  Length
        10.  Sequence
        11.  Edge weight
        12.  ProbFolding
        13.  Log-Probfolding
        14.  Tf
        15. maxTm
        16. maxSeq
        17. has14
        18. dGtotal
        19. dGintrin
        20. dGloop
        21. dGconc

        '''

        # Check if the oligo group exists
        # Doesn't exist if the clustering is not performed

        oligo_group_key = -1
        if self.current_break.oligo_group:
            oligo_group_key = self.current_break.oligo_group.key

        cvs_row =      ['.'.join([str(x) for x in self.current_break.oligo.key]),
                        oligo_group_key,
                        '.'.join([str(x) for x in self.current_break.key]),
                        self.current_break.type,
                        self.current_break.location,
                        '.'.join([str(x) for x in self.next_break.key]),
                        self.next_break.type,
                        self.next_break.location,
                        self.edge_length,
                        ''.join([str(x) for x in self.ssDNA_seq_list]),
                        self.edge_weight,
                        self.edge_prob,
                        self.edge_logprob,
                        self.edge_Tf,
                        self.edge_maxTm,
                        self.edge_maxseq,
                        int(self.edge_has14),
                        self.dG_total,
                        np.sum(self.dG_intrin_list),
                        np.sum(self.dG_inter_list),
                        self.dG_conc]

        return cvs_row

    def set_edge_weight(self):
        '''
        Set edge weight
        '''
        self.edge_weight = self.origami.autobreak.optimize(self)

    def is_valid(self):
        '''
        Determine if edge is valid
        '''
        return self.valid and self.active and (not self.current_break.dont_break and
                                               not self.next_break.dont_break and
                                               not self.current_break.dont_break_temp and
                                               not self.next_break.dont_break_temp)

    def make_connection(self, from_break, to_break):
        '''
        Make the connection between two break nodes
        '''

        # Set the break nodes
        self.current_break = from_break
        self.next_break    = to_break

        # Assign the edge weights
        self.update_connection()

    def update_connection(self):
        '''
        Update edge weights
        '''

        # Get the temperature parameter for dG optimization
        temperature_kelvin = self.autobreak.optim_temperature_kelvin

        # Initialize sequence and dna list
        self.sequence_list  = []
        self.ssDNA_seq_list = []
        self.dsDNA_seq_list = []

        self.ssDNA_pos_list = []
        self.dsDNA_pos_list = []

        # Make the sequence list
        self.sequence_list.append(self.current_break.sequence)

        # Check if the breaks are in consecutive positions on the same sequence
        if not (self.current_break.sequence == self.next_break.sequence and
                self.current_break.direction*(self.next_break.idx - self.current_break.idx) > 0):

            # Get forward link
            next_sequence = self.current_break.sequence.next_sequence

            # Iterate over all the sequences
            while next_sequence:
                self.sequence_list.append(next_sequence)

                # Stop if the sequence is same as the final sequence
                if next_sequence == self.next_break.sequence:
                    break

                # Update next sequence
                next_sequence = next_sequence.next_sequence

        # Iterate over sequence list
        if len(self.sequence_list) == 1:
            start_point   = self.current_break.break_point_adjusted+1
            final_point   = self.next_break.break_point_adjusted+1
            dna_sequence  = self.current_break.strand.dna[start_point:final_point]
            position_list = self.current_break.strand.scaffoldPos[start_point:final_point]

            if len(position_list) > 0:
                self.ssDNA_pos_list.append(position_list)
            if len(dna_sequence) > 0:
                self.ssDNA_seq_list.append(dna_sequence)
        else:
            # 1. Get the 5' sequence
            start_point   = self.current_break.break_point_adjusted+1
            final_point   = self.current_break.sequence.strHigh+1
            dna_sequence  = self.current_break.strand.dna[start_point:final_point]
            position_list = self.current_break.strand.scaffoldPos[start_point:final_point]

            if len(position_list) > 0:
                self.ssDNA_pos_list.append(position_list)
            if len(dna_sequence) > 0:
                self.ssDNA_seq_list.append(dna_sequence)

            # 2. Get the sequences in between
            for sequence in self.sequence_list[1:-1]:
                self.ssDNA_pos_list.append(sequence.scaffoldPos)
                self.ssDNA_seq_list.append(sequence.dna)

            # 3. Get the 3' sequence
            start_point   = self.next_break.sequence.strLow
            final_point   = self.next_break.break_point_adjusted+1
            dna_sequence  = self.next_break.strand.dna[start_point:final_point]
            position_list = self.next_break.strand.scaffoldPos[start_point:final_point]

            if len(position_list) > 0:
                self.ssDNA_pos_list.append(position_list)
            if len(dna_sequence) > 0:
                self.ssDNA_seq_list.append(dna_sequence)

        # Remove empty sequences
        self.dsDNA_seq_list = [dna.strip() for dna in self.ssDNA_seq_list if len(dna.strip()) > 0]
        self.dsDNA_pos_list = [list(filter(lambda x: x, pos_list)) for pos_list in self.ssDNA_pos_list]

        # Replace all empty characters in ssDNA seq list with ?
        self.ssDNA_seq_list = [dna.replace(" ", "?") for dna in self.ssDNA_seq_list]

        # Remove empty lists from positions list
        self.dsDNA_pos_list = list(filter(lambda x: len(x), self.dsDNA_pos_list))

        # Determine mean values for the positions
        self.dsDNA_mean_pos_list = [np.round(np.mean(x)) for x in self.dsDNA_pos_list]

        # Determine Tm
        self.Tm_list       = np.array([utilities.sequence_to_Tm(dna) for dna in self.dsDNA_seq_list])

        # 1. Determine intrinsic free energies
        self.dG_intrin_list = []
        self.dH_intrin_list = []
        self.dS_intrin_list = []

        for dna in self.dsDNA_seq_list:
            (dGintrin,
             dHintrin,
             dSintrin) = utilities.sequence_to_dG(dna, temperature_kelvin)

            # Add free energies to the lists
            self.dG_intrin_list.append(dGintrin)
            self.dH_intrin_list.append(dHintrin)
            self.dS_intrin_list.append(dSintrin)

        # Make the lists arrays
        self.dG_intrin_list = np.array(self.dG_intrin_list)
        self.dH_intrin_list = np.array(self.dH_intrin_list)
        self.dS_intrin_list = np.array(self.dS_intrin_list)

        # Get scaffold length
        scaffold_length   = self.origami.oligos['scaffold'][0].length
        scaffold_circular = self.origami.oligos['scaffold'][0].circular

        # 2. Determine interfacial coupling energies
        self.dG_inter_list = []
        self.dS_inter_list = []

        for i in range(len(self.dsDNA_mean_pos_list)-1):
            dGinter, dSinter = utilities.position_to_loop_dG(self.dsDNA_mean_pos_list[i],
                                                             self.dsDNA_mean_pos_list[i+1],
                                                             scaffold_length, scaffold_circular,
                                                             temperature_kelvin)
            self.dG_inter_list.append(dGinter)
            self.dS_inter_list.append(dSinter)

        # Make the lists arrays
        self.dG_inter_list = np.array(self.dG_inter_list)
        self.dS_inter_list = np.array(self.dS_inter_list)

        # 3. Get free energy due to concentration
        dGconc, dSconc = utilities.conc_to_dG(temperature_kelvin)

        self.dG_conc = dGconc
        self.dS_conc = dSconc

        # Get RT values
        self.RT = 0.593/298.15*temperature_kelvin

        # 4. Get total energies
        self.dG_total = np.sum(self.dG_intrin_list) + np.sum(self.dG_inter_list) + self.dG_conc
        self.dS_total = np.sum(self.dS_intrin_list) + np.sum(self.dS_inter_list) + self.dS_conc
        self.dH_total = np.sum(self.dH_intrin_list)

        # 5. Determine probabilities
        self.edge_prob = np.exp(-self.dG_total/self.RT)/(1.0+np.exp(-self.dG_total/self.RT))

        # 6. Determine log-probabilities
        self.edge_logprob = np.log(self.edge_prob)

        # 7. Determine Tf in Celcius
        self.edge_Tf = self.dH_total/self.dS_total - 273.15

        # Determine lengths
        self.ssDNA_length_list = np.array([len(dna) for dna in self.ssDNA_seq_list])
        self.dsDNA_length_list = np.array([len(dna) for dna in self.dsDNA_seq_list])

        # Determine the edge weights
        if len(self.Tm_list) > 0:
            self.edge_maxTm = max(self.Tm_list)
        else:
            self.edge_maxTm = 0

        # Length parameters
        if len(self.dsDNA_length_list) > 0:
            self.edge_maxseq = max(self.dsDNA_length_list)
        else:
            self.edge_maxseq = 0

        self.edge_num14  = np.sum(self.dsDNA_length_list >= 14)
        self.edge_has14  = int(self.edge_num14 > 0)

        self.edge_num16  = np.sum(self.dsDNA_length_list >= 16)
        self.edge_has16  = int(self.edge_num16 > 0)

        # Tm parameters
        self.edge_numTm  = np.sum(self.Tm_list >= self.LOW_TM)
        self.edge_hasTm  = np.sum(self.edge_numTm > 0)

        # Determine structure score
        self.edge_num_cross = len(self.dsDNA_seq_list)
        self.edge_structure = self.edge_num_cross**2.0

        # Set edge weight
        self.set_edge_weight()

        # Check dsDNA list to decide on the validity of edge
        if len(self.dsDNA_length_list) == 0:
            self.valid = False

    def make_loop_edge(self):
        '''
        Make loop edge
        '''
        if self.current_break == self.next_break:
            self.isloop                  = True
            self.current_break.loop_edge = self


class BreakPath:
    def __init__(self, break_node, break_edge=None, score=0):
        '''
        Break path object
        '''
        self.break_node = break_node
        self.break_edge = break_edge
        self.score      = score


class BreakNode:
    def __init__(self):
        '''
        Break node class
        '''
        self.crossover        = None
        self.next_break       = None
        self.previous_break   = None
        self.neighbor_break   = None
        self.connected_breaks = None
        self.break_edges      = None
        self.edge_nodes       = None
        self.type             = None
        self.sequence         = None
        self.crossover        = None
        self.loop_edge        = None

        # Nucleotide parameters
        self.current_nucleotide = None
        self.next_nucleotide    = None

        # State parameters
        self.visited          = False
        self.active           = True
        self.break_state      = None   # broken for break,  not-broken for no break
        self.dont_break       = False  # If set to True, keep the break not-broken
        self.dont_break_temp  = False  # Transient version of dont_break. If set to True, keep the break not-broken

        # Cluster info
        self.oligo_group      = None

        # Graph parameter
        self.score               = -utilities.INFINITY
        self.best_path_node      = None
        self.best_path_nodes     = None
        self.shortest_paths      = None
        self.k_potential_paths   = None
        self.k_shortest_paths    = None
        self.traverse_path       = None
        self.shortest_score      = 0
        self.order_id            = None

    def is_dsDNA(self):
        '''
        Determine if the break node is located on dsDNA
        '''
        # Initialize dsDNA parameters
        current_dsDNA = True
        next_dsDNA = True

        if self.current_nucleotide:
            current_dsDNA = self.current_nucleotide.dsDNA
        else:
            current_dsDNA = False

        if self.next_nucleotide:
            next_dsDNA = self.next_nucleotide.dsDNA
        else:
            next_dsDNA = False

        return current_dsDNA and next_dsDNA

    def reset_break_path(self):
        self.order_id         = -1
        self.traverse_path    = []
        self.best_path_nodes  = []
        self.best_path_node   = None
        self.score            = -utilities.INFINITY
        self.visited          = False
        self.shortest_paths   = []

    def reset_k_paths(self):
        self.k_shortest_paths  = []
        self.k_potential_paths = []

    def is_break_edge_possible(self, other_break):
        '''
        Check if break edge is possible from self to to another break node for a circular oligo
        '''
        order_difference = self.get_break_order_difference(other_break)

        # Criteria for a proper connection (no loop) in the right direction
        return order_difference > 0 or (not order_difference == 0 and order_difference == -self.order_id)

    def get_loop_edge(self):
        '''
        Return loop edge
        '''

        return self.loop_edge

    def get_break_distance(self, other_break):
        '''
        Break to break distance
        '''
        distance = other_break.distance - self.distance
        if distance <= 0 and self.oligo.circular:
            return distance+self.oligo.length
        else:
            return distance

    def get_break_order_difference(self, other_break):
        '''
        Return break to order id difference
        '''
        return other_break.order_id - self.order_id

    def get_valid_edge_nodes(self):
        '''
        Get break nodes connected by edges
        '''
        # Get the connected break nodes
        return [break_edge.next_break for break_edge in self.break_edges if not break_edge.next_break.dont_break]

    def get_valid_edges(self):
        '''
        Get edges that lead to break nodes that can be broken
        '''

        # Get the connected break nodes
        return [break_edge for break_edge in self.break_edges if break_edge.is_valid()]

    def get_k_shortest_paths(self, final_break, k_num=10, k_select='best'):
        '''
        Get k-shortest path results
        '''
        # Initialize k-shortest paths
        self.k_shortest_paths = []

        # 1. Get the shortest paths
        shortest_path = self.get_shortest_path(final_break)

        # If the here is no path found return empty list
        if shortest_path is None:
            return self.k_shortest_paths

        # 2.Add best path to k-path list
        self.k_shortest_paths  = [shortest_path]
        self.k_potential_paths = []
        num_k_solutions        = 1

        # Check the final score of the path
        if shortest_path.score == 0:
            return self.k_shortest_paths

        # 3. Make the paths
        while num_k_solutions < k_num:

            # Get the last best path
            last_solution = self.k_shortest_paths[-1]

            # Iterate through the edges
            for i in range(len(last_solution.edges[:-1])):
                # Inactive edge list
                inactive_edges = []

                for j in range(len(self.k_shortest_paths)):

                    if last_solution.is_identical(self.k_shortest_paths[j], i):
                        inactive_edges.append(last_solution.edges[i])

                        # Make the edge inactive
                        last_solution.edges[i].active = False

                # Reset break paths and find solution
                self.oligo.reset_break_paths()
                self.oligo.reset_break_order_ids(self, final_break)

                # Get sub solution
                potential_solution = self.get_shortest_path(final_break)

                # Make the edges active again
                for edge in inactive_edges:
                    edge.active = True

                # If there is no solution continue
                if potential_solution is None:
                    continue

                # Get the first solution in the list
                new_solution = potential_solution

                # Add potential solution if it doesnt exist in k-shortest paths
                potential_solution_exists = False
                for solution in self.k_shortest_paths:
                    if solution.is_identical(new_solution):
                        potential_solution_exists = True
                        break

                # Add new solution to potential paths
                if not potential_solution_exists:
                    self.k_potential_paths.append(new_solution)

            # Check if the potential path list is empty, if empty quit
            if len(self.k_potential_paths) == 0:
                break

            # Sort the potential paths and add the best one to k-shortest paths list
            self.k_potential_paths.sort(key=lambda x: x.score, reverse=True)

            if k_select == 'best':
                solution_index = 0
            else:
                # Add random one to the k-shortest paths
                solution_index = random.randint(0, len(self.k_potential_paths)-1)

            # Add item to k-shortest path list
            self.k_shortest_paths.append(self.k_potential_paths[solution_index])

            # Update num solutions
            num_k_solutions += 1

            # Remove the result from potential paths list
            self.k_potential_paths.pop(solution_index)

        return self.k_shortest_paths

    def get_shortest_path(self, final_break):
        '''
        Find the shortest path between current and final break points
        '''
        # Initialize the set and stack
        stack = [self]

        while stack:
            # Pop the break node
            new_break = stack.pop(0)

            # If current node is final break, quit
            if final_break.visited and new_break == final_break:
                break

            # Get valid break edges
            valid_break_edges = new_break.get_valid_edges()

            # Update the scores for connected breaks
            for break_edge in valid_break_edges:
                # If id difference is in wrong direction and if it is a loop, discard the edge
                if not new_break.is_break_edge_possible(break_edge.next_break):
                    continue

                # Determine the new score
                if new_break.order_id == 0:
                    new_score = break_edge.edge_weight
                else:
                    new_score = new_break.score + break_edge.edge_weight

                # Update the score based on the existence of a neighbor crossover
                if new_break.neighbor_break in new_break.best_path_nodes:
                    new_score += -utilities.INFINITY*utilities.INFINITY

                # Make new break Path object
                new_break_path = BreakPath(new_break, break_edge, new_score)

                # If new score is higher than the previous one, make a new list
                if new_score > break_edge.next_break.score:
                    break_edge.next_break.best_path_node  = new_break_path
                    break_edge.next_break.best_path_nodes = new_break.best_path_nodes + [new_break]
                    break_edge.next_break.score           = new_score

                # Add next break to connected breaks list
                if not break_edge.next_break.visited:
                    stack.append(break_edge.next_break)

                # Make the next break visited
                break_edge.next_break.visited = True

        # Finally compare with the loop connection
        if self == final_break and self.loop_edge and self.loop_edge.is_valid():
            # Make a loop break path object
            loop_break_path = BreakPath(self, self.loop_edge, self.loop_edge.edge_weight)

            if self.loop_edge.edge_weight > final_break.score:
                final_break.best_path_node = loop_break_path
                final_break.score          =  self.loop_edge.edge_weight

        # Return best path
        final_break.shortest_path = final_break.traverse_best_path(self)

        return final_break.shortest_path

    def traverse_best_path(self, start_break):

        # Initialize shortest path
        self.shortest_path = None

        # Check if the final break has a path node
        if self.best_path_node is None:
            return self.shortest_path

        # Make the traverse path for starting node empty
        self.traverse_path = []

        # Make final node break path object
        final_break_path = BreakPath(self, None, self.score)

        # Initalize the traverse path for the initial node
        self.best_path_node.break_node.traverse_path = self.traverse_path + [final_break_path]

        # Assign new break path
        new_break_path = self.best_path_node

        # Assign new break
        new_break = new_break_path.break_node

        while new_break:

            if new_break == start_break:
                new_break_solution              = OligoBreakSolution()
                new_break_solution.start_break  = start_break
                new_break_solution.final_break  = self
                new_break_solution.break_paths  = new_break.traverse_path + [new_break_path]
                new_break_solution.score        = self.score

                # Initialize the solution
                new_break_solution.initialize()

                self.shortest_path = new_break_solution
                break

            # Get the new path nodes
            new_path_node = new_break.best_path_node

            # Check if there is a path from new break in reverse direction
            if new_path_node is None:
                break

            # Iterate through each node
            new_path_node.break_node.traverse_path = new_break.traverse_path + [new_break_path]

            # Update new break path
            new_break_path = new_path_node

            # Update new break
            new_break = new_path_node.break_node

        return self.shortest_path

    def get_connected_breaks(self):
        '''
        Return connected breaks
        '''
        # Initialize connected breaks
        self.connected_breaks = []

        # 1. Add the directly connected next break
        if self.next_break:
            self.connected_breaks.append(self.next_break)

        # 2. Add the directly connected previous break
        if self.previous_break:
            self.connected_breaks.append(self.previous_break)

        # 3. Add the neighbor break
        if self.neighbor_break:
            self.connected_breaks.append(self.neighbor_break)

        return self.connected_breaks

    def break_cadnano(self):
        '''
        Break the breaknode
        '''
        if self.type == 'crossover':
            self.origami.remove_cadnano_crossover(self.vh, self.idx, self.direction)
        else:
            self.origami.split_cadnano_strand(self.vh, self.idx, self.direction)

    def depth_first_search(self):
        '''
        Depth-first-search graph traverse algorithm to find connected components
        '''
        # Initialize the set and stack
        visited, stack = set(), [self]
        while stack:
            # Pop the break node
            new_break = stack.pop()
            if new_break not in visited:
                # Add break to visited list
                visited.add(new_break)

                # Make the break visited
                new_break.visited = True

                # Get the connected breaks
                connected_breaks = new_break.get_connected_breaks()

                # Extend the group with new breaks
                stack.extend(set(connected_breaks) - visited)

        return list(visited)


def main():

    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("-i",   "--input",    type=str,
                        help="Cadnano json file")

    parser.add_argument("-readonly",   "--readonly",  action='store_true',
                        help="Read-only to determine oligo scores")

    parser.add_argument("-rule",   "--rule",     type=str, default='xstap',
                        help="Break rule")

    parser.add_argument("-score",   "--score",     type=str, default='sum',
                        help="Optimization score function")

    parser.add_argument("-func",   "--func",     type=str, default='dG:50',
                        help="Optimization function")

    parser.add_argument("-seq",   "--sequence", type=str, default=None,
                        help="Sequence file in txt")

    parser.add_argument("-sort", "--sort", action='store_true',
                        help="Sort oligos for stepwise optimization.")

    parser.add_argument("-nsol",   "--nsol",     type=int,
                        help="Number of solutions", default=10)

    parser.add_argument("-minlength",   "--minlength",     type=int,
                        help="Minimum staple length", default=21)

    parser.add_argument("-dontb",   "--dontbreak",  type=int,
                        help="Dont break oligos less than the length specified", default=0)

    parser.add_argument("-v",   "--verbose",  action='store_true',
                        help="Verbose output")

    parser.add_argument("-permute",   "--permute",  action='store_true',
                        help="Permute sequence")

    parser.add_argument("-npermute",  "--npermute",  type=int,
                        help="Number of permutation iterations", default=0)

    parser.add_argument("-writeall",   "--writeall",  action='store_true',
                        help="Write all results")

    parser.add_argument("-seed",   "--seed",  type=int, default=0,
                        help="Random seed")

    args = parser.parse_args()

    # Check if the required arguments are passed to the code
    if args.input is None:
        parser.print_help()
        sys.exit('Input file does not exist!')

    # Assign the parameters
    input_filename          = args.input
    sequence_filename       = args.sequence
    read_only               = args.readonly
    break_rule              = utilities.parse_break_rule(args.rule)
    score_func              = utilities.parse_score_function(args.score)
    optimization_func       = utilities.parse_optim_function(args.func)
    global_solutions        = args.nsol
    dontbreak_less_than     = args.dontbreak
    verbose_output          = args.verbose
    random_seed             = args.seed
    permute_sequence        = args.permute
    write_all_results       = args.writeall
    shuffle_oligos          = not args.sort
    npermute                = args.npermute

    # Create args dictionary
    args_dict = {'input': args.input,
                 'sequence': args.sequence,
                 'readonly': args.readonly,
                 'rule': args.rule,
                 'score': args.score,
                 'func': args.func,
                 'nsol': args.nsol,
                 'minlength': args.minlength,
                 'dontb': args.dontbreak,
                 'verbose': args.verbose,
                 'seed': args.seed,
                 'permute': args.permute,
                 'npermute': args.npermute,
                 'writeall': args.writeall,
                 'sort': args.sort}

    # Check if input file exists
    if not os.path.isfile(input_filename):
        sys.exit('Input file does not exist!')

    # Set random seed in order to get the same results with the same set of parameters
    random.seed(random_seed)

    # Create Origami object
    new_origami = origamidesign.Origami()

    # Create Autobreak object and make cross assignments
    new_autobreak = AutoBreak()

    # Store args
    new_autobreak.args_dict = args_dict

    # Cross assignments
    new_origami.autobreak = new_autobreak
    new_autobreak.origami = new_origami

    # Assign break rule
    new_autobreak.set_break_rule(break_rule)

    # Set solution numbers
    new_autobreak.set_solution_nums(1, global_solutions)

    # Set optimization  functions
    new_autobreak.set_optimization_func(optimization_func)

    # Set score functions
    new_autobreak.set_score_func(score_func)

    # Set permute sequence
    new_autobreak.set_permute_sequence(permute_sequence)

    # Set oligo shuffle parameter
    new_autobreak.set_oligo_shuffle_parameter(shuffle_oligos)

    # Preprocess optimization parameters
    new_autobreak.preprocess_optim_params()

    # Set verbose parameter
    new_autobreak.set_verbose_output(verbose_output)

    # Set output directory
    new_autobreak.set_output_directory(input_filename)

    # Set write all flag
    new_autobreak.set_write_all_results(write_all_results)

    # Set optimization temperature
    new_autobreak.set_temperature_parameter()

    # Set lower bound for oligo lengths
    new_autobreak.set_lower_bound(args.minlength)

    # Set read only setting
    new_autobreak.set_readonly(read_only)

    # Initialize origami object
    new_origami.initialize(input_filename)

    # Set sequence filename
    new_origami.set_sequence_file(sequence_filename)

    # Set circularize
    new_origami.set_circularize(True)

    # Define json output
    new_autobreak.define_output_files()

    # Write input arguments
    new_autobreak.write_input_args()

    # Check if the scaffold is circular
    new_origami.warn_circular_scaffold()

    # Check if it is a read-only or autobreak run
    if read_only:

        # Prepare origami
        new_origami.prepare_origami()

        # Run the permutation protocol
        new_autobreak.permute_scaffold_sequence_readonly(npermute)

        # Correct sequence offsets
        new_autobreak.correct_complete_solution_offsets()

        # Compare complete solutions
        new_autobreak.compare_complete_solutions()

        # Write results summary
        new_autobreak.write_results_summary()

        # Set best result's sequence offset
        new_autobreak.set_best_sequence_offset()

    else:
        # Prepare origami for autobreak
        new_origami.prepare_origami()

        # Cluster staples
        new_origami.cluster_oligo_groups()

        # Set dont break flag for oligos
        new_origami.set_dont_break_oligos(dontbreak_less_than)

        # Make break-break graph
        new_autobreak.initialize()

        # Create results excel file
        new_autobreak.create_results_excel_file()

        # Run the permutation protocol
        new_autobreak.permute_scaffold_sequence_autobreak(npermute)

        # Correct sequence offsets
        new_autobreak.correct_complete_solution_offsets()

        # Compare complete solutions
        new_autobreak.compare_complete_solutions()

        # Write results summary
        new_autobreak.write_results_summary()

        # Write best result
        new_autobreak.write_best_result()

        # Break best solutions
        new_autobreak.break_best_complete_solution()

        # Set best result's sequence offset
        new_autobreak.set_best_sequence_offset()

    # Read only the scores
    new_autobreak.determine_oligo_scores()

    # Color oligos by folding prob
    new_autobreak.color_oligos_by_Tf()

    # Export initial scores
    new_autobreak.export_initial_scores()

    # Split scaffold
    new_origami.split_scaffold()

    # Write result to json
    new_autobreak.write_final_part_to_json()

    # Write sequence file to output directory
    new_autobreak.copy_sequence_file()


if __name__ == "__main__":
    main()
