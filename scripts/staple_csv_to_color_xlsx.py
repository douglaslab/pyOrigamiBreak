# This script converts a csv staple file exported from Cadnano into an Excel xlsx format.
# It stylizes the font and fill color of the hex codes in the Color column to make them easier to interpret.
# We recommend running in a virtual environment after installing the dependencies with this command
#   pip install pandas openpyxl

import sys

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

class ColorFormatter:

    def is_color_dark(self, hex_color):
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        luminance = (0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2])
        return luminance < 128

    def convert_rgb_to_argb(self, rgb_color):
        if rgb_color.startswith("#"):
            rgb_color = rgb_color[1:]  # Remove the "#" if present
        return "FF" + rgb_color.upper()  # Add alpha channel (FF for full opacity)

    def apply_formatting_to_all_sheets(self, workbook_path, column_name):
        wb = openpyxl.load_workbook(workbook_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            column_letter = None
            for col in sheet.iter_cols(1, sheet.max_column):
                if col[0].value == column_name:
                    column_letter = col[0].column_letter
                    break
            if column_letter:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet[f'{column_letter}{row}']
                    cell_value = cell.value
                    if cell_value and cell_value.startswith("#"):
                        argb_color = self.convert_rgb_to_argb(cell_value)
                        cell.fill = PatternFill(start_color=argb_color, end_color=argb_color, fill_type="solid")
                        font_color = "FFFFFF" if self.is_color_dark(cell_value) else "000000"
                        cell.font = Font(color=font_color)

        wb.save(workbook_path)

if __name__ == "__main__":

    formatter = ColorFormatter()

    if len(sys.argv) != 2:
        print("Usage: python staple_csv_to_color_xlsx.py <filename>")
        sys.exit(1)
    csv_file = sys.argv[1]
    excel_file = csv_file.replace(".csv", ".xlsx")  # Add suffix to output file
    column_name = 'Color' # input("Enter the name of the column containing hex colors: ")
    try:
        df = pd.read_csv(csv_file)  # Read CSV into DataFrame
        df.to_excel(excel_file, index=False)  # Convert to Excel
        formatter.apply_formatting_to_all_sheets(excel_file, column_name)
        print(f"Excel file created and formatted at: {excel_file}")
    except FileNotFoundError:
        print(f"Error: File not found - {csv_file}")
    except Exception as e:
        print(f"An error occurred: {e}")

