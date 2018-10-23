#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Date    : 2018-03-03 15:44:55
# @Author  : Your Name (you@example.org)
# @Link    : http://example.org
# @Version : $Id$

from cadnano.document import Document
from shutil import copyfile

import os
import csv
import sys
import cadnano
import argparse
import glob
import utilities
import openpyxl
import collections
import numpy as np


# GLOBALS
WHITE_FONT = openpyxl.styles.Font(color=openpyxl.styles.colors.WHITE)


class Project:
    '''
    Origami Project class that keeps multiple structures
    '''
    def __init__(self):
        self.json_files     = []
        self.structures     = []
        self.plates_96well  = []
        self.plates_384well = []
        self.date           = ''
        self.author         = ''
        self.stocks         = {}
        self.oligos_list    = []
        self.oligos_dict    = {}

        self.stock_counts   = {}    # Number of stocks
        self.stock_keys     = []
        self.total_bases    = 0     # Total number of bases
        self.num_structures = 0
        self.wb_96well      = None  # 96  well plate workbook object
        self.wb_384well     = None  # 384 well plate workbook object

        self.echo_input     = []
        self.echo_res_input = []

        self.noskip         = False
        self.nreps_ECHO     = 1

    def set_nreps_ECHO(self, nreps=1):
        '''
        Set nreps_ECHO
        '''
        self.nreps_ECHO = nreps

    def set_noskip(self, no_skip=False):
        '''
        Set noskip parameter
        '''
        self.noskip = no_skip

    def assign_color_counters(self):
        '''
        Assign color counters
        '''

        for key in self.oligos_dict:
            self.oligos_dict[key].colorctr = self.color_counts[self.oligos_dict[key].color]

    def assign_sortkeys(self):
        '''
        Assign sorkeys
        '''
        for key in self.oligos_dict:
            self.oligos_dict[key].sortkey  = (self.oligos_dict[key].colorctr,
                                              self.oligos_dict[key].color,
                                              self.oligos_dict[key].bitseq)

    def assign_bitseqs(self):
        '''
        Assign bitseqs
        '''
        for key in self.oligos_dict:
            self.oligos_dict[key].bitseq   = ''.join([str(x) for x in self.oligos_dict[key].bitlist])

    def prepare_oligos_list(self):
        '''
        Prepare oligos list
        '''
        self.oligos_list = []
        for key in self.oligos_dict:

            # Add staple to list
            self.oligos_list.append(self.oligos_dict[key])

        # 4. Sort staples based on length
        self.oligos_list.sort(key=lambda x: (x.length, x.vh5p, x.idx5p, x.vh3p, x.idx3p, x.sequence))

        # 5. Sort staples based on the sort key
        self.oligos_list.sort(key=lambda x: x.sortkey, reverse=True)

    def count_colors(self):
        '''
        Count colors
        '''
        self.color_counts = {}
        for key in self.oligos_dict:
            color = self.oligos_dict[key].color
            if color not in self.color_counts:
                self.color_counts[color] = 0
            self.color_counts[color] += 1

    def add_structure(self, new_structure):
        '''
        Add new structure
        '''
        self.structures.append(new_structure)

        # Reset bits for the new structure
        new_structure.reset_bits(self.num_structures)

        # Add staples to main list
        for key in new_structure.oligos_dict:

            if key not in self.oligos_dict:
                self.oligos_dict[key] = new_structure.oligos_dict[key]
            else:
                new_structure.oligos_dict[key] = self.oligos_dict[key]

        # Assign project
        new_structure.project = self

    def assign_bit_id(self, new_structure):
        '''
        Add new staples to staples list
        '''

        # Assign the bit for the staples
        for key in self.oligos_dict:
            if key in new_structure.oligos_dict:
                self.oligos_dict[key].bitlist[new_structure.structure_id] = 1

    def read_sequence(self, sequence_file):
        '''
        Read sequence file
        '''

        # 1. Check the popular scaffolds list
        # 2. Check if the file exists

        # Initialize scaffold sequence
        self.scaffold_sequence = ''

        if sequence_file in utilities.SCAFFOLD_SEQUENCES:
            self.scaffold_sequence = utilities.SCAFFOLD_SEQUENCES[sequence_file]

        elif os.path.isfile(sequence_file):
            # Read sequence from file
            f = open(sequence_file)
            self.scaffold_sequence = ''.join([line.strip() for line in f.readlines()])
            f.close()

            # Convert to upper case
            self.scaffold_sequence = self.scaffold_sequence.upper()

        return self.scaffold_sequence

    def parse_input_files(self, input_fnames):
        '''
        Parse input files
        '''

        self.json_files = []
        for input_fname in input_fnames:
            self.json_files += glob.glob(input_fname)

        # Filter out any non-json file
        self.json_files = list(filter(lambda x: os.path.splitext(x)[1] == '.json', self.json_files))

        # Sort json files
        self.json_files.sort()

        return self.json_files

    def _create_workbook_96well(self):
        '''
        Create workbook for 96well plate output format
        '''
        self.wb_96well = openpyxl.Workbook()

    def _create_workbook_384well(self):
        '''
        Create workbook for 384 well output format
        '''
        self.wb_384well = openpyxl.Workbook()

    def calculate_bases(self):
        '''
        Calculate total number of bases
        '''
        self.total_bases = 0

        for current_oligo in self.oligos_list:
            # Update total bases
            self.total_bases += current_oligo.length

    def prepare_stocks(self):
        '''
        Prepare stocks
        '''
        self.stocks       = {}
        self.stock_counts = {}
        self.stock_keys   = []

        # Initialize stock id
        stock_id = 0

        for current_oligo in self.oligos_list:

            if current_oligo.sortkey not in self.stock_keys:
                # Create new stock
                new_stock = Stock()
                new_stock.key         = current_oligo.sortkey
                new_stock.count       = 0
                new_stock.stock_id    = stock_id
                new_stock.oligos_list = [current_oligo]
                new_stock.color       = current_oligo.color

                self.stocks[current_oligo.sortkey]       = new_stock
                self.stock_counts[current_oligo.sortkey] = 0
                self.stock_keys.append(current_oligo.sortkey)

                # Update stock id
                stock_id += 1

            # Update oligo counts for each stock
            self.stock_counts[current_oligo.sortkey] += 1
            self.stocks[current_oligo.sortkey].count += 1
            self.stocks[current_oligo.sortkey].oligos_list.append(current_oligo)

    def set_output_directory(self, output_directory=None):
        '''
        Set output directory
        '''
        # Split first input file
        head, tail       = os.path.split(os.path.abspath(self.json_files[0]))
        root, ext        = os.path.splitext(tail)

        # Save the filename with head removed only tail
        self.input_tail     = tail

        # List existing output directories
        potential_directories = list(filter(lambda x: os.path.isdir(x),
                                            glob.glob(head+'/'+root+'_autopipette_[0-9][0-9][0-9]')))

        # Get the number extensions
        number_extensions = [int(x[-3:]) for x in potential_directories]

        # Get the counter
        output_counter = 1
        if len(number_extensions) > 0:
            output_counter = max(number_extensions)+1

        # Check the output directory
        if output_directory is None:
            self.output_directory = head+'/'+root+"_autopipette_%03d" % (output_counter)
        else:
            self.output_directory = output_directory

        # Make the output directory
        os.makedirs(self.output_directory, exist_ok=True)

        # Copy input files to output directory
        for input_file in self.json_files:
            head, tail = os.path.split(os.path.abspath(input_file))
            copyfile(input_file, self.output_directory+'/'+tail)

    def add_json_files(self, json_files):
        '''
        Add json files
        '''
        self.json_files = json_files

        # Get number of structures
        self.num_structures = len(self.json_files)

    def write_ECHO_input(self, out_fname):
        '''
        Write ECHO input
        '''

        with open(out_fname, 'w') as csvfile:

            # Create cvs writer
            self.echo_writer = csv.writer(csvfile, delimiter=',')

            # Write header
            header = ['Source Plate Name',
                      'Source Well',
                      'Destination Plate Name',
                      'Destination Well',
                      'Transfer Volume']

            self.echo_writer.writerow(header)

            # Iterate over each row
            for echo_row in self.echo_res_input+self.echo_input:
                echo_list = [echo_row['sourcePlate'],
                             echo_row['sourceWell'],
                             echo_row['destPlate'],
                             echo_row['destWell'],
                             echo_row['volume']]
                self.echo_writer.writerow(echo_list)

    def prepare_ECHO_reservoir_input(self):
        '''
        Prepare ECHO reservoir input
        '''
        self.echo_res_input = []

        # Prepare echo input for each structure
        for current_structure in self.structures:
            current_structure.prepare_ECHO_reservoir_input()

        # Iterate over destination plates
        for dest_plate in self.plates_dest:

            # Iterate over each well in destination plate
            ordered_structures = collections.OrderedDict(sorted(dest_plate.structures_dict.items()))

            for well_id, current_structure in ordered_structures.items():

                dest_echo_input = {'destPlate': dest_plate.plate_label,
                                   'destWell': well_id}

                # Iterate over all the echo input entries for a structure
                for echo_row in current_structure.echo_res_input:
                    self.echo_res_input.append({**dest_echo_input, **echo_row})

        # Sort the input
        self.sort_ECHO_reservoir_input()

    def sort_ECHO_input(self):
        '''
        Sort echo input by source and dest plates
        '''
        # Sort by source well
        self.echo_input = sorted(self.echo_input, key=lambda entry: (entry['sourceWell'][0],
                                                                     int(entry['sourceWell'][1:])))
        # Sort by source plate
        self.echo_input = sorted(self.echo_input, key=lambda entry: entry['sourcePlate'])

        # Sort by destination plate
        self.echo_input = sorted(self.echo_input, key=lambda entry: entry['destPlate'])

    def sort_ECHO_reservoir_input(self):
        '''
        Sort ECHO reservoir input
        '''
        self.echo_res_input = sorted(self.echo_res_input, key=lambda entry: (entry['destWell'][0],
                                                                             int(entry['destWell'][1:])))
        # Sort by source plate
        self.echo_res_input = sorted(self.echo_res_input, key=lambda entry: entry['destPlate'])

    def prepare_ECHO_input(self):
        '''
        Prepare ECHO input
        '''

        # Prepare echo input for each structure
        for current_structure in self.structures:
            current_structure.prepare_ECHO_input()

        # Echo input
        self.echo_input = []

        # Iterate over destination plates
        for dest_plate in self.plates_dest:

            # Iterate over each well in destination plate
            ordered_structures = collections.OrderedDict(sorted(dest_plate.structures_dict.items()))

            for well_id, current_structure in ordered_structures.items():

                dest_echo_input = {'destPlate': dest_plate.plate_label,
                                   'destWell': well_id}

                # Iterate over all the echo input entries for a structure
                for echo_row in current_structure.echo_input:
                    self.echo_input.append({**dest_echo_input, **echo_row})

        # Sort echo input
        self.sort_ECHO_input()

    def _prepare_structure_sheet_96well(self):
        '''
        Prepare structures
        '''
        for i in range(self.num_structures):
            # Get structure
            current_structure = self.structures[i]

            # Stock row
            current_structure.stock_row = [self.json_files[i]]

            # Remove json extension
            structure_name, file_ext = os.path.splitext(self.json_files[i])

            # Assign structure name
            current_structure.structure_name = structure_name

            # Starting H2O volume
            current_structure.water_96well = 300

            # Stock keys for the structure
            current_structure.stock_keys = []
            for j in range(len(self.stock_keys)):
                # Get current stock key
                current_stock_key = self.stock_keys[j]

                # Get membership key
                membership_key = current_stock_key[2]

                # Check stock membership
                if membership_key[i] == '1':
                    # Add stock key to structure
                    current_structure.stock_keys.append(current_stock_key)

                    # Stock row
                    current_structure.stock_row.append('%d (%d)' % (j+1, self.stocks[current_stock_key].count))

                    # Water update
                    current_structure.water_96well -= self.stocks[current_stock_key].count

                    # Add stock to structure
                    current_structure.stocks[current_stock_key] = self.stocks[current_stock_key]

            # Add water
            current_structure.stock_row.append('H2O (%d)' % (current_structure.water_96well))

            # Add to structure worksheet
            self.ws_str.append(current_structure.stock_row)

    def _color_structure_sheet_96well(self):
        '''
        Color structrue sheet for 96 well plate
        '''
        # Get maximum row and column for structrue sheet
        maxRow = self.ws_str.max_row
        maxCol = self.ws_str.max_column

        for colNum in range(2, maxCol + 1, 1):
            for rowNum in range(2, maxRow + 1):
                # Get cell
                current_cell = self.ws_str.cell(row=rowNum, column=colNum)
                if current_cell.value and current_cell.value[0] != 'H':

                    # Get stock id and stock color
                    stock_id    = int(current_cell.value.split('(')[0]) - 1
                    stock_color = openpyxl.styles.colors.Color(rgb=self.stock_keys[stock_id][1][1:])

                    # Set color
                    current_cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=stock_color)

                    # Set font
                    current_cell.font = WHITE_FONT

    def _write_structure_sheet_96well(self):
        '''
        Write structure sheet for 96 well plate format output
        '''
        self.ws_str = self.wb_96well.active
        self.ws_str.title = "Structures"

        # Update column width
        self.ws_str.column_dimensions['A'].width = 20

        # Add header row
        self.ws_str.append(('Structure', 'Stocks (ul)'))

        # Calculate bases and stocks
        self.calculate_bases()
        self.prepare_stocks()

        # Prepare structures sheet
        self._prepare_structure_sheet_96well()

        # Color cells for structures sheet
        self._color_structure_sheet_96well()

        # Append total base count
        self.ws_str.append(['Total bases', self.total_bases])

    def _write_structure_sheet_384well(self, nreps=2):
        '''
        Write structure sheet for 384 well
        '''

        # Delete the first sheet
        self.wb_384well.remove(self.wb_384well.active)

        # Initialize counter plate
        counter_plate = Plate()
        counter_plate.current_row_id   = 0
        counter_plate.current_plate_id = 0

        # Initialize current plate
        current_plate = None
        ws_current    = None

        # Get plate label
        self.sheet_labels = []

        # Destination plates
        self.plates_dest = []

        # Create a new destination plate
        for i in range(len(self.structures)):
            current_structure = self.structures[i]

            # Set rwo and column id
            counter_plate.current_row_id  += 1
            counter_plate.current_col_id   = 0

            for j in range(nreps):

                # Advance in row order
                counter_plate.advance_row_order()

                # Plate label
                plate_label       = 'DestPlate-%d' % (counter_plate.current_plate_id)

                # Check if plate is in the list
                if plate_label not in self.sheet_labels:

                    # Add plate label to sheet label list

                    self.sheet_labels.append(plate_label)
                    ws_current = self.wb_384well.create_sheet(title=plate_label)

                    # Create new 96well plate
                    current_plate = Plate()
                    current_plate.structures_dict  = {}
                    current_plate.structures_list  = []
                    current_plate.worksheet        = ws_current
                    current_plate.plate_label      = plate_label
                    current_plate.plate_id         = counter_plate.current_plate_id
                    current_plate.set_dimensions()

                    # Add current plate to plates list
                    self.plates_dest.append(current_plate)

                # Get well id
                well_id = counter_plate.get_current_well_id()

                # Assign stucture value
                ws_current[well_id] = current_structure.structure_name

                # Current plate
                current_plate.structures_dict[well_id] = current_structure
                current_plate.structures_list.append(current_structure)

    def _write_plate_sheets_384well(self, plate_header=''):
        '''
        Write plate sheets for ECHO mode - 384 well plate order
        '''
        # Header for the plate sheets
        self.plate_column_header = ['Plate Name', 'Well Position', 'Sequence Name', 'Sequence',
                                    'Oligo Start', 'Oligo End', 'Oligo Length', 'Oligo Color',
                                    'Oligo Membership']

        # Create a dummy counter plate
        counter_plate = Plate()
        counter_plate.initialize()
        counter_plate.set_dimensions(plate_size=384)
        counter_plate.current_col_id = 0

        # Aggregate sheet
        self.ws_ALL = self.wb_384well.create_sheet(title='PlatesAll')

        # Add header row for the new sheet
        self.ws_ALL.append(self.plate_column_header)

        # Set column widths
        self.ws_ALL.column_dimensions['A'].width = 20
        self.ws_ALL.column_dimensions['D'].width = 60
        self.ws_ALL.column_dimensions['I'].width = 20

        # Initialize the current plate
        current_plate = None

        for current_oligo in self.oligos_list:

            # Advance row order
            counter_plate.advance_row_order()

            # Get plate id
            plate_label = 'Plate%s-%d' % (self.plate_header, counter_plate.current_plate_id)

            # Get well id
            well_id  = counter_plate.get_current_well_id()

            # Check if plate is in the list
            if plate_label not in self.sheet_labels:

                self.sheet_labels.append(plate_label)
                ws_current = self.wb_384well.create_sheet(title=plate_label)

                # Add header row for the new sheet
                ws_current.append(self.plate_column_header)

                # Set column widths
                ws_current.column_dimensions['A'].width = 20
                ws_current.column_dimensions['D'].width = 60
                ws_current.column_dimensions['I'].width = 20

                # Create new 96well plate
                current_plate = Plate()
                current_plate.oligos_dict = {}
                current_plate.oligos_list = []
                current_plate.worksheet   = ws_current
                current_plate.plate_label = plate_label
                current_plate.plate_id    = counter_plate.current_plate_id
                current_plate.set_dimensions(plate_size=384)

                # Add current plate to plates list
                self.plates_384well.append(current_plate)

            # Row
            current_plate.oligos_dict[well_id]    = current_oligo
            current_plate.oligos_list.append(current_oligo)

            # Update current oligo entries
            current_oligo.plate384_plate_label = plate_label
            current_oligo.plate384_well_id     = well_id
            current_oligo.plate384_seq_id      = current_oligo.plate96_seq_id
            current_oligo.plate384             = current_plate

            # 384 well plate row for excel sheet
            current_oligo._make_plate384_row()

            # Write the row
            ws_current.append(current_oligo.plate384_row)

            # Add row to all plates sheet
            self.ws_ALL.append(current_oligo.plate384_row)

    def _write_plate_sheets_96well(self, plate_header=''):

        # Header for the plate sheets
        self.plate_column_header = ['Plate Name', 'Well Position', 'Sequence Name', 'Sequence',
                                    'Stock Id', 'Oligo Start', 'Oligo End',
                                    'Oligo Length', 'Oligo Color', 'Oligo Membership']

        # Create a dummy counter plate
        counter_plate = Plate()
        counter_plate.initialize()
        counter_plate.current_col_id = 0

        # Initialize sequence and stock id
        seq_id    = 0
        stock_id  = 1

        # Stock id/key
        prev_stock_key    = self.oligos_list[0].sortkey
        current_stock_key = None

        # Store sheet labels
        self.sheet_labels = [self.ws_str.title]

        # Save plate header
        self.plate_header = plate_header

        # Get plate label and assign the first startwell
        plate_label = 'Plate%s-%d' % (self.plate_header, counter_plate.current_plate_id)
        self.stocks[prev_stock_key].startwell = (plate_label, 'A1')

        # Aggregate sheet
        self.ws_ALL = self.wb_96well.create_sheet(title='PlatesAll')

        # Add header row for the new sheet
        self.ws_ALL.append(self.plate_column_header)

        # Set column widths
        self.ws_ALL.column_dimensions['A'].width = 20
        self.ws_ALL.column_dimensions['D'].width = 60
        self.ws_ALL.column_dimensions['J'].width = 20

        # Initialize the current plate
        current_plate = None

        for current_oligo in self.oligos_list:
            # Get current stock key
            current_stock_key = current_oligo.sortkey

            # Update sequence id
            seq_id += 1

            if current_stock_key == prev_stock_key or self.noskip:
                counter_plate.advance_row_order()
            else:
                counter_plate.advance_stock_row_order()

            # Update stock id
            if current_stock_key != prev_stock_key:
                stock_id += 1

            # Get plate id
            plate_label = 'Plate%s-%d' % (self.plate_header, counter_plate.current_plate_id)

            # Get well id
            well_id  = counter_plate.get_current_well_id()

            # Update start and endwells
            if current_stock_key == prev_stock_key:
                self.stocks[current_stock_key].endwell   = (plate_label, well_id)
            else:
                self.stocks[current_stock_key].startwell = (plate_label, well_id)
                self.stocks[current_stock_key].endwell   = (plate_label, well_id)

            # Check if plate is in the list
            if plate_label not in self.sheet_labels:

                self.sheet_labels.append(plate_label)
                ws_current = self.wb_96well.create_sheet(title=plate_label)

                # Add header row for the new sheet
                ws_current.append(self.plate_column_header)

                # Set column widths
                ws_current.column_dimensions['A'].width = 20
                ws_current.column_dimensions['D'].width = 60
                ws_current.column_dimensions['J'].width = 20

                # Create new 96well plate
                current_plate = Plate()
                current_plate.oligos_dict = {}
                current_plate.oligos_list = []
                current_plate.worksheet   = ws_current
                current_plate.plate_label = plate_label
                current_plate.plate_id    = counter_plate.current_plate_id

                # Add current plate to plates list
                self.plates_96well.append(current_plate)

            # Row
            current_plate.oligos_dict[well_id] = current_oligo
            current_plate.oligos_list.append(current_oligo)

            # Update current oligo entries
            current_oligo.plate96_plate_label = plate_label
            current_oligo.plate96_well_id     = well_id
            current_oligo.plate96_seq_id      = seq_id
            current_oligo.plate96             = current_plate
            current_oligo.stock               = self.stocks[current_stock_key]
            current_oligo.stock_id            = stock_id

            # 96 well plate row for excel sheet
            current_oligo._make_plate96_row()

            # Write the row
            ws_current.append(current_oligo.plate96_row)

            # Add row to all plates sheet
            self.ws_ALL.append(current_oligo.plate96_row)

            # Update stock keys
            prev_stock_key = current_stock_key

    def _color_structures_sheet_96well(self):
        # Color cells for structures sheet
        for current_sheet in self.wb_96well:
            if not current_sheet.title[:5] == 'Plate':
                continue
            maxRow = current_sheet.max_row
            for rowNum in range(2, maxRow + 1):
                current_cell = current_sheet.cell(row=rowNum, column=5)
                if current_cell.value:
                    stock_id    = int(current_cell.value) - 1
                    stock_color = openpyxl.styles.colors.Color(rgb=self.stock_keys[stock_id][1][1:])

                    # Update fill color and font
                    current_cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=stock_color)
                    current_cell.font = WHITE_FONT

    def _write_stocks_sheet_96well(self):
        # Write stocks sheet
        self.ws_stocks = self.wb_96well.create_sheet(title='Stocks')
        # Append header
        self.ws_stocks.append(['Stock Id', 'Start Plate', 'Start Well', 'End Plate', 'End Well', '#Oligos'])

        for current_stock_key in self.stock_keys:
            self.ws_stocks.append([self.stocks[current_stock_key].stock_id+1,
                                   self.stocks[current_stock_key].startwell[0],
                                   self.stocks[current_stock_key].startwell[1],
                                   self.stocks[current_stock_key].endwell[0],
                                   self.stocks[current_stock_key].endwell[1],
                                   self.stocks[current_stock_key].count])

        # Change the colors
        maxRow = self.ws_stocks.max_row
        for rowNum in range(2, maxRow + 1):
            current_cell = self.ws_stocks.cell(row=rowNum, column=1)
            if current_cell.value:
                stock_id    = int(current_cell.value) - 1
                stock_color = openpyxl.styles.colors.Color(rgb=self.stock_keys[stock_id][1][1:])
                current_cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=stock_color)
                current_cell.font = WHITE_FONT

    def _order_sheets_96well(self):
        # Order the sheets

        sheet_order = np.arange(len(self.wb_96well._sheets), dtype=int)
        sheet_order[2:] -= 1
        sheet_order[1] = len(self.wb_96well._sheets) - 1
        self.wb_96well._sheets = [self.wb_96well._sheets[i] for i in sheet_order]

    def save_sheets_96well(self, out_fname):
        # Save excel file (workbook)
        self.wb_96well.save(out_fname)

    def save_sheets_384well(self, out_fname):
        # Save excel file (workbook)
        self.wb_384well.save(out_fname)

    def write_oligos_96well(self, out_fname, plate_header=''):
        '''
        Export oligos
        '''
        self._create_workbook_96well()
        self._write_structure_sheet_96well()
        self._write_plate_sheets_96well(plate_header)
        self._color_structures_sheet_96well()
        self._write_stocks_sheet_96well()
        self._order_sheets_96well()
        self.save_sheets_96well(out_fname)

    def write_oligos_384well(self, out_fname, plate_header=''):
        '''
        Export oligos
        '''
        self._create_workbook_384well()
        self._write_structure_sheet_384well(nreps=self.nreps_ECHO)
        self._write_plate_sheets_384well(plate_header)
        self.save_sheets_384well(out_fname)


class Structure:
    '''
    Structure class that keeps the oligos to order and the Cadnano input file
    '''
    def __init__(self):
        self.project        = None
        self.structure_name = ''
        self.structure_id   = ''
        self.json_file      = ''
        self.oligos_list    = []
        self.oligos_dict    = {}
        self.project        = None
        self.stocks         = {}
        self.water_96well   = 300          # In ul (microliters)
        self.water_384well  = 20000        # In nl (nanoliters)
        self.echo_drop_vol  = 50           # In nl (nanoliters)
        self.echo_input     = []
        self.echo_res_input = []           # Echo reservoir input
        self.oligo_conc     = 200          # In uM

    def prepare_oligos_list(self):
        '''
        Prepare oligos list
        '''
        self.oligos_list = []
        for key, oligo in self.oligos_dict.items():
            if oligo is not None:
                self.oligos_list.append(oligo)

    def sort_oligos_list(self):
        '''
        Sort oligos list
        '''
        self.oligos_list = sorted(self.oligos_list, key=lambda oligo: oligo.sortkey)

    def prepare_ECHO_input(self):
        '''
        Prepare echo input
        '''
        self.echo_input = []

        # Prepare oligos list and sort
        self.prepare_oligos_list()
        self.sort_oligos_list()

        for current_oligo in self.oligos_list:
            # Get plate id
            plate_label = current_oligo.plate384.plate_label

            # Well id
            well_id     = current_oligo.plate384_well_id

            # Volume
            volume      = self.echo_drop_vol

            # Prepare echo input
            self.echo_input.append({'sourcePlate': plate_label,
                                    'sourceWell': well_id,
                                    'volume': volume})

    def get_echo_input(self):
        '''
        Get echo input
        '''
        return self.echo_input

    def prepare_ECHO_reservoir_input(self):
        '''
        Prepare ECHO reservoir input
        '''
        self.water_384well  = 20000.0 - 1.0*len(self.oligos_list)*self.echo_drop_vol

        # Prepare reservoir input
        self.echo_res_input = [{'sourcePlate': 'Reservoir',
                                'sourceWell': 'A1',
                                'volume': self.water_384well}]

    def reset_bits(self, num_structures):
        '''
        Reset bits
        '''
        for key in self.oligos_dict:
            self.oligos_dict[key].bitlist = [0]*num_structures

    def read_oligos(self, cadnano_part, scaffold_sequence, add_T=False):
        self.cadnano_oligos = cadnano_part.oligos()
        self.cadnano_oligos = sorted(self.cadnano_oligos, key=lambda x: x.length(), reverse=True)

        # Apply sequence to scaffold
        self.cadnano_oligos[0].applySequence(scaffold_sequence)

        # Initialize the scaffolds and staples
        self.oligos_dict = {}

        # Empty character
        empty_ch = '?'
        if add_T:
            empty_ch = 'T'

        # Iterate over oligos
        for oligo in self.cadnano_oligos[1:]:
            strand5p  = oligo.strand5p()
            strand3p  = oligo.strand3p()

            # Make new oligo
            new_oligo           = Oligo()
            new_oligo.length    = oligo.length()
            new_oligo.vh5p      = strand5p.idNum()
            new_oligo.vh3p      = strand3p.idNum()
            new_oligo.idx5p     = strand5p.idx5Prime()
            new_oligo.idx3p     = strand3p.idx3Prime()
            new_oligo.color     = oligo.getColor()
            new_oligo.sequence  = ''

            # Get the strand generator
            generator           = oligo.strand5p().generator3pStrand()
            for strand in generator:
                if strand.totalLength() != len(strand.sequence()):
                    new_oligo.sequence += strand.length()*empty_ch
                else:
                    new_oligo.sequence += strand.sequence().replace(' ', empty_ch)

            new_oligo.key       = '-'.join([str(new_oligo.vh5p),
                                            str(new_oligo.idx5p),
                                            str(new_oligo.vh3p),
                                            str(new_oligo.idx3p),
                                            new_oligo.sequence])

            new_oligo.startkey  = '%d[%d]' % (new_oligo.vh5p, new_oligo.idx5p)
            new_oligo.finishkey = '%d[%d]' % (new_oligo.vh3p, new_oligo.idx3p)

            self.oligos_dict[new_oligo.key] = new_oligo

        return self.oligos_dict


class Plate:
    '''
    Plate class
    '''
    def __init__(self):
        self.plate_label = ''
        self.plate_id    = 0
        self.plate_size  = 96
        self.oligos_dict = {}
        self.oligos_list = []
        self.structures  = []
        self.worksheet   = None

        self.row_length  = 8
        self.col_length  = 12

        # Transient parameters
        self.current_plate_id = 0
        self.current_row_id   = 0
        self.current_col_id   = 0
        self.current_well_id  = ''

    def initialize(self):
        self.current_plate_id = 1
        self.current_row_id   = 1
        self.current_col_id   = 1

    def set_dimensions(self, plate_size=96):
        # Assign plate dimensions
        self.plate_size = plate_size

        if self.plate_size == 384:
            self.row_length = 16
            self.col_length = 24
        elif self.plate_size == 96:
            self.row_length = 8
            self.col_length = 12

    def advance_stock_row_order(self):
        '''
        Advance stock row order
        '''
        # Set col id to 1
        self.current_col_id = 1

        if self.current_row_id == self.row_length:
            self.current_row_id    = 1
            self.current_plate_id += 1
        else:
            self.current_row_id = (self.current_row_id % self.row_length) + 1

    def advance_stock_col_order(self):
        '''
        Advance stock col order
        '''
        # Set row id to 1
        self.current_row_id = 1

        # Check new row id
        if self.current_col_id == self.col_length:
            self.current_col_id    = 1
            self.current_plate_id += 1
        else:
            self.current_col_id = (self.current_col_id % self.col_length) + 1

    def advance_row_order(self):
        '''
        Advance row order
        '''
        # Get new row id
        if self.current_col_id == self.col_length:
            self.current_row_id += 1

        # Get new col id
        self.current_col_id = (self.current_col_id % self.col_length) + 1

        # Check new row id
        if self.current_row_id > self.row_length:
            self.current_row_id    = 1
            self.current_plate_id += 1

    def n_advance_row_order(self, nsteps=12):
        '''
        Advance n-steps in row order
        '''
        for i in range(nsteps):
            self.advance_row_order()

    def n_advance_col_order(self, nsteps=8):
        '''
        Advance n-steps in col order
        '''
        for i in range(nsteps):
            self.advance_col_order()

    def advance_col_order(self):
        '''
        Advance col order
        '''
        # Get new column id
        if self.current_row_id == self.row_length:
            self.current_col_id += 1

        # Get new col id
        self.current_row_id = (self.current_row_id % self.row_length) + 1

        # Check new row id
        if self.current_col_id > self.col_length:
            self.current_col_id    = 1
            self.current_plate_id += 1

    def get_current_well_id(self):
        '''
        Get current well id
        '''
        self.current_well_id = chr(64+self.current_row_id)+str(self.current_col_id)

        return self.current_well_id

    def get_current_row_col(self):
        '''
        Get current row-col
        '''

        return (self.current_row_id, self.current_col_id)

    def get_current_plate_id(self):
        '''
        Get current plate id
        '''
        return self.current_plate_id


class Stock:
    '''
    Origami stock class
    '''
    def __init__(self):
        self.key         = ''
        self.color       = ''
        self.count       = 0
        self.stock_id    = ''
        self.stock_name  = ''
        self.oligos_list = []
        self.oligos_dict = {}

        self.startwell   = ()
        self.endwell     = ()


class Oligo:
    def __init__(self):
        self.key      = None
        self.color    = None
        self.colorctr = None
        self.sequence = None
        self.vh5p     = None
        self.idx5p    = None
        self.vh3p     = None
        self.idx3p    = None
        self.bitlist  = None
        self.bitseq   = None
        self.sortkey  = None

        # Plate and stock parameters
        self.plate96_well_id  = None
        self.plate96_seq_id   = None
        self.plate96          = None

        self.plate384_well_id = None
        self.plate384_seq_id  = None
        self.plate384         = None

        self.stock            = None

    def _make_plate96_row(self):
        '''
        Make 96 well plate row for the oligo
        '''
        self.plate96_row = [self.plate96_plate_label,
                            self.plate96_well_id,
                            self.plate96_seq_id,
                            self.sequence,
                            self.stock_id,
                            self.startkey,
                            self.finishkey,
                            self.length,
                            self.color,
                            self.bitseq]

    def _make_plate384_row(self):
        '''
        Make 384 well plate row for the oligo
        '''
        self.plate384_row = [self.plate384_plate_label,
                             self.plate384_well_id,
                             self.plate384_seq_id,
                             self.sequence,
                             self.startkey,
                             self.finishkey,
                             self.length,
                             self.color,
                             self.bitseq]


def parse_input_files(input_fnames):
    '''
    Parse input files
    '''

    json_files = []
    for input_fname in input_fnames:
        json_files += glob.glob(input_fname)

    # Filter out any non-json file
    json_files = list(filter(lambda x: os.path.splitext(x)[1] == '.json', json_files))

    # Sort json files
    json_files.sort()

    return json_files


def main():

    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument("-i",   "--input",    type=str, nargs='+', required=True,
                        help="Cadnano json file(s)")

    parser.add_argument("-o",   "--output",    type=str,
                        help="Output directory", default=None)

    parser.add_argument("-seq",   "--seq", type=str,
                        help="Scaffold sequence file")

    parser.add_argument("-nreps", "--nreps", type=int, default=1,
                        help="Number of replicates for ECHO output")

    parser.add_argument("-header", "--header", type=str, default='',
                        help="Plate header")

    parser.add_argument("-offset", "--offset", type=int, default=None,
                        help="Sequence offset")

    parser.add_argument("-addT", "--addT", action='store_true',
                        help="Replace ? with T")

    parser.add_argument("-noskip", "--noskip", action='store_true',
                        help="Do not skip wells for 96 well plate format")

    args = parser.parse_args()

    # Check if the required arguments are passed to the code
    if args.input is None:
        parser.print_help()
        sys.exit('Input file does not exist!')

    # Get json files
    json_files = parse_input_files(args.input)

    # Check if the json file list is empty
    if len(json_files) == 0:
        sys.exit('Input file does not exist!')

    # Create a project
    new_project = Project()

    # Get scaffold sequence
    scaffold_sequence      = new_project.read_sequence(args.seq)

    # Add json files
    new_project.add_json_files(json_files)

    # Set output directory
    new_project.set_output_directory(args.output)

    # Set no skip
    new_project.set_noskip(args.noskip)

    # Set Echo replicates
    new_project.set_nreps_ECHO(args.nreps)

    # Check if sequence file exists
    if len(scaffold_sequence) == 0:
        sys.exit('Scaffold sequence is not available!')

    # Define output file
    xlsx_output_96well  = new_project.output_directory+'/oligos_96well.xlsx'
    xlsx_output_384well = new_project.output_directory+'/oligos_384well.xlsx'
    echo_output_384well = new_project.output_directory+'/echo_384well.csv'

    # Initialize cadnano
    app = cadnano.app()
    doc = app.document = Document()

    # 1. Iterate over all the json files
    for i in range(len(json_files)):

        # Get json file
        json_input = json_files[i]

        # Define json output
        head, ext  = os.path.splitext(os.path.abspath(json_input))

        # Read cadnano input file
        doc.readFile(json_input)

        # Get part
        part = doc.activePart()

        # Get staples
        # Create structure object
        new_structure = Structure()
        new_structure.offset = args.offset or part.getSequenceOffset()

        # Apply the offset
        new_structure.scaffold_sequence = (new_project.scaffold_sequence[new_structure.offset:] +
                                           new_project.scaffold_sequence[:new_structure.offset])
        new_structure.oligos_dict  = new_structure.read_oligos(part, new_structure.scaffold_sequence, args.addT)
        new_structure.structure_id = i
        new_structure.project      = new_project

        # Add structure to list
        new_project.add_structure(new_structure)

        # Assign the bit for the staples
        new_project.assign_bit_id(new_structure)

    # 2. Count colors
    new_project.count_colors()

    # 3. Assign the color counters for the staples and make bitseq
    new_project.assign_color_counters()
    new_project.assign_bitseqs()
    new_project.assign_sortkeys()
    new_project.prepare_oligos_list()

    # 7. Export oligos for 96 well plate format
    new_project.write_oligos_96well(xlsx_output_96well, args.header)

    # 8. Export oligos for 384 well plate format
    new_project.write_oligos_384well(xlsx_output_384well, args.header)

    # 9. Prepare ECHO input
    new_project.prepare_ECHO_input()

    # 10. Prepare ECHO reservoir input
    new_project.prepare_ECHO_reservoir_input()

    # 11. Write ECHO inputs
    new_project.write_ECHO_input(echo_output_384well)


if __name__ == "__main__":
    main()
